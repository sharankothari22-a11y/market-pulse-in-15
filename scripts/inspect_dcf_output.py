#!/usr/bin/env python3
"""
Inspect a DCF_Output_*.xlsm produced by DCF_Multi_Source_Pipeline_REFACTORED.ipynb.

Loads the workbook twice (formulas + cached values), prints every sheet's
dimensions and first 10 non-empty cells, locates the DCF-relevant labels
(Fair Value, Revenue, FCF, WACC, Terminal Value), and emits a final verdict:

    REAL_OUTPUT    — numeric DCF values and formulas with cached results
    EMPTY_TEMPLATE — no valuation numbers, just template labels
    PARTIAL        — some values present but key DCF outputs missing/blank

Usage:
    python scripts/inspect_dcf_output.py [path/to/DCF_Output_*.xlsm]

If no path given, uses the newest notebooks/DCF_Output_*.xlsm.
"""
from __future__ import annotations

import glob
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


REPO_ROOT = Path(__file__).resolve().parent.parent
NOTEBOOKS_DIR = REPO_ROOT / "notebooks"

LABELS_OF_INTEREST = [
    "fair value",
    "revenue",
    "projections",
    "fcf",
    "free cash flow",
    "dcf",
    "terminal value",
    "wacc",
    "implied share price",
    "share price",
]


def find_newest_output() -> Path | None:
    hits = sorted(
        glob.glob(str(NOTEBOOKS_DIR / "DCF_Output_*.xlsm")),
        key=lambda p: Path(p).stat().st_mtime,
        reverse=True,
    )
    return Path(hits[0]) if hits else None


def fmt_val(v: Any) -> str:
    if v is None:
        return "<none>"
    if isinstance(v, float):
        return f"{v:,.4f}"
    s = str(v)
    return s if len(s) <= 60 else s[:57] + "..."


def iter_non_empty(ws, limit: int | None = None):
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is None or cell.value == "":
                continue
            yield cell
            count += 1
            if limit and count >= limit:
                return


def inspect(xlsm_path: Path) -> str:
    print(f"\n{'=' * 78}")
    print(f"INSPECTING: {xlsm_path}")
    print(f"Size: {xlsm_path.stat().st_size:,} bytes")
    print(f"{'=' * 78}")

    wb_f = load_workbook(xlsm_path, keep_vba=True, data_only=False)
    wb_v = load_workbook(xlsm_path, keep_vba=True, data_only=True)

    print(f"\nSheets ({len(wb_f.sheetnames)}):")
    for name in wb_f.sheetnames:
        ws = wb_f[name]
        print(f"  - '{name}'  {ws.max_row} x {ws.max_column}")

    # Collect stats for verdict
    total_cells = 0
    numeric_cells = 0
    formula_cells = 0
    formula_cached = 0
    labels_found: dict[str, list[tuple[str, str, Any, Any]]] = {
        lbl: [] for lbl in LABELS_OF_INTEREST
    }

    for sheet_name in wb_f.sheetnames:
        ws_f = wb_f[sheet_name]
        ws_v = wb_v[sheet_name]

        print(f"\n{'-' * 78}\nSheet: '{sheet_name}'")
        print(f"{'-' * 78}")

        # First 10 non-empty cells
        print("First 10 non-empty cells:")
        for cell in iter_non_empty(ws_f, limit=10):
            coord = cell.coordinate
            val = cell.value
            cached = ws_v[coord].value
            is_formula = isinstance(val, str) and val.startswith("=")
            suffix = ""
            if is_formula:
                suffix = f"  [cached={fmt_val(cached)}]"
            print(f"  {coord:>6}  {fmt_val(val)}{suffix}")

        # Full scan for stats + labels
        for row in ws_f.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                v = cell.value
                if v is None or v == "":
                    continue
                total_cells += 1
                if isinstance(v, (int, float)):
                    numeric_cells += 1
                elif isinstance(v, str) and v.startswith("="):
                    formula_cells += 1
                    cached = ws_v[cell.coordinate].value
                    if cached is not None and cached != "" and cached != 0:
                        formula_cached += 1
                elif isinstance(v, str):
                    lv = v.strip().lower()
                    for lbl in LABELS_OF_INTEREST:
                        if lbl in lv:
                            # Look at the cell to the right (common layout:
                            # label in col N, value in col N+1).
                            right = ws_f.cell(
                                row=cell.row, column=cell.column + 1
                            ).value
                            right_cached = ws_v.cell(
                                row=cell.row, column=cell.column + 1
                            ).value
                            labels_found[lbl].append(
                                (sheet_name, cell.coordinate, v, right_cached)
                            )
                            break

    # Labels report
    print(f"\n{'=' * 78}\nLABELS OF INTEREST\n{'=' * 78}")
    for lbl, hits in labels_found.items():
        if not hits:
            print(f"  [{lbl:<22}] no match")
            continue
        print(f"  [{lbl:<22}] {len(hits)} match(es)")
        for sheet, coord, label_text, right_val in hits[:3]:
            print(
                f"    - {sheet}!{coord}: '{fmt_val(label_text)}' → "
                f"next col = {fmt_val(right_val)}"
            )

    # Verdict
    print(f"\n{'=' * 78}\nSTATS\n{'=' * 78}")
    print(f"  total non-empty cells : {total_cells:,}")
    print(f"  numeric cells         : {numeric_cells:,}")
    print(f"  formula cells         : {formula_cells:,}")
    print(f"    of which cached     : {formula_cached:,}")

    has_key_labels = any(
        labels_found[k]
        for k in ("fair value", "wacc", "terminal value", "fcf", "revenue")
    )
    has_real_numbers = numeric_cells >= 50 or formula_cached >= 20

    if has_real_numbers and has_key_labels:
        verdict = "REAL_OUTPUT"
        reason = (
            f"{numeric_cells} numeric cells, {formula_cached} formulas with "
            f"cached values, key DCF labels present"
        )
    elif has_key_labels and not has_real_numbers:
        verdict = "EMPTY_TEMPLATE"
        reason = (
            "labels exist but very few numeric/cached values — looks like the "
            "template without computed output"
        )
    elif has_real_numbers and not has_key_labels:
        verdict = "PARTIAL"
        reason = "numbers present but key DCF labels missing"
    else:
        verdict = "EMPTY_TEMPLATE"
        reason = "neither numeric output nor key DCF labels present"

    print(f"\n{'=' * 78}\nVERDICT: {verdict}\n  {reason}\n{'=' * 78}")
    return verdict


def main():
    if len(sys.argv) > 1:
        path = Path(sys.argv[1]).resolve()
    else:
        path = find_newest_output()
        if path is None:
            print(
                f"ERROR: no DCF_Output_*.xlsm found in {NOTEBOOKS_DIR}",
                file=sys.stderr,
            )
            sys.exit(2)

    if not path.exists():
        print(f"ERROR: file not found: {path}", file=sys.stderr)
        sys.exit(2)

    verdict = inspect(path)
    sys.exit(0 if verdict == "REAL_OUTPUT" else 1)


if __name__ == "__main__":
    main()
