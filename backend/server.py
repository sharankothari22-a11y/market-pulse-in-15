"""
Market Pulse — Hardened Backup Backend
=======================================
Philosophy: NEVER 500. NEVER crash. ALWAYS return usable JSON.

Every endpoint:
  1. Tries the real data source (yfinance, RSS, etc.)
  2. Falls back to last-known-good cache in MongoDB
  3. Falls back to safe default structure
  4. Returns with a `stale: true` or `fallback: true` flag so the frontend
     can show a subtle indicator without breaking the UI.

Failure modes explicitly handled:
  - Missing env vars (MongoDB URL, API keys) → safe defaults, log warning
  - yfinance rate-limited / returning empty → cached prices, then mock structure
  - Network timeouts on any external call → cached, then mock
  - MongoDB unavailable → in-memory cache for the request
  - Bad ticker / unknown symbol → structured error response, not 500
  - Any exception in any endpoint body → structured 200 with fallback data

Replace backend/server.py with this file. Keep the existing imports of
supporting modules (collectors/, research_platform/) intact — this file
uses the same collector function names so the file is drop-in compatible.
"""

from __future__ import annotations

# ---------- Standard library ----------
import os
import sys
import json
import re
import logging
import asyncio
import traceback
import subprocess
import functools
import dataclasses
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import Any, Optional

# ---------- Third-party (all imports are defensive) ----------
try:
    from fastapi import FastAPI, APIRouter, HTTPException, Request, UploadFile, File
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.responses import JSONResponse, FileResponse
    from pydantic import BaseModel
except Exception as e:
    # If FastAPI itself fails to import, we can't serve anything. Raise loudly.
    raise RuntimeError(f"FATAL: FastAPI import failed: {e}")

# Optional imports — every one is wrapped so startup never crashes.
try:
    from dotenv import load_dotenv
    ROOT_DIR = Path(__file__).parent
    load_dotenv(ROOT_DIR / ".env")
except Exception as e:
    ROOT_DIR = Path(__file__).parent
    print(f"[startup] dotenv skipped: {e}")

try:
    from motor.motor_asyncio import AsyncIOMotorClient
    MOTOR_AVAILABLE = True
except Exception as e:
    MOTOR_AVAILABLE = False
    print(f"[startup] motor not available: {e}")

try:
    import yfinance as yf
    YF_AVAILABLE = True
except Exception as e:
    YF_AVAILABLE = False
    yf = None
    print(f"[startup] yfinance not available: {e}")

try:
    import httpx
    HTTPX_AVAILABLE = True
except Exception as e:
    HTTPX_AVAILABLE = False
    print(f"[startup] httpx not available: {e}")

# ---------- research_platform — the real research engine ----------
# Each import is wrapped so one broken module doesn't disable the others.
# If ANY import fails, the corresponding feature falls back to simple logic.
RP_AVAILABLE = False
_rp_errors = []
try:
    # Ensure research_platform is on the path (it's a sibling folder)
    _rp_path = Path(__file__).parent / "research_platform"
    if _rp_path.exists() and str(_rp_path) not in sys.path:
        sys.path.insert(0, str(_rp_path))
except Exception as e:
    _rp_errors.append(f"path_setup: {e}")

try:
    from ai_engine.session_manager import new_session as rp_new_session, load_session as rp_load_session, list_sessions as rp_list_sessions
    RP_SESSION_MGR = True
except Exception as e:
    RP_SESSION_MGR = False
    rp_new_session = rp_load_session = rp_list_sessions = None
    _rp_errors.append(f"session_manager: {e}")

try:
    from ai_engine.scenario_engine import run_scenarios as rp_run_scenarios
    RP_SCENARIOS = True
except Exception as e:
    RP_SCENARIOS = False
    rp_run_scenarios = None
    _rp_errors.append(f"scenario_engine: {e}")

try:
    from ai_engine.scoring import score_session as rp_score_session
    RP_SCORING = True
except Exception as e:
    RP_SCORING = False
    rp_score_session = None
    _rp_errors.append(f"scoring: {e}")

try:
    from ai_engine.pdf_builder import build_report as rp_build_report
    RP_PDF = True
except Exception as e:
    RP_PDF = False
    rp_build_report = None
    _rp_errors.append(f"pdf_builder: {e}")

try:
    from ai_engine.audit_export import export_to_excel as rp_export_excel, export_to_html as rp_export_html
    RP_EXPORT = True
except Exception as e:
    RP_EXPORT = False
    rp_export_excel = rp_export_html = None
    _rp_errors.append(f"audit_export: {e}")

try:
    from ai_engine.swot import generate_swot as rp_generate_swot
    RP_SWOT = True
except Exception as e:
    RP_SWOT = False
    rp_generate_swot = None
    _rp_errors.append(f"swot: {e}")

try:
    from ai_engine.porter import generate_porter as rp_generate_porter
    RP_PORTER = True
except Exception as e:
    RP_PORTER = False
    rp_generate_porter = None
    _rp_errors.append(f"porter: {e}")

try:
    from ai_engine.signal_detector import detect_signals as rp_detect_signals
    RP_SIGNALS = True
except Exception as e:
    RP_SIGNALS = False
    rp_detect_signals = None
    _rp_errors.append(f"signal_detector: {e}")

try:
    from ai_engine.factor_engine import signals_to_factors as rp_signals_to_factors
    RP_FACTORS = True
except Exception as e:
    RP_FACTORS = False
    rp_signals_to_factors = None
    _rp_errors.append(f"factor_engine: {e}")

RP_AVAILABLE = RP_SESSION_MGR and RP_SCENARIOS and RP_SCORING
if RP_AVAILABLE:
    print(f"[startup] ✓ research_platform loaded (scenarios, scoring, sessions)")
else:
    print(f"[startup] ⚠ research_platform partial/unavailable: {_rp_errors}")

# =====================================================================
# LOGGING
# =====================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger("market_pulse")

# =====================================================================
# CONFIG — every value has a safe default
# =====================================================================

MONGO_URL = os.environ.get("MONGO_URL", "").strip()
DB_NAME = os.environ.get("DB_NAME", "market_pulse").strip()
CORS_ORIGINS = os.environ.get("CORS_ORIGINS", "*").split(",")
TWELVE_DATA_KEY = os.environ.get("TWELVE_DATA_KEY", "").strip()
FINNHUB_KEY = os.environ.get("FINNHUB_API_KEY", "").strip()
ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY", "").strip()

# Cache TTL for "stale-but-useful" data (seconds)
CACHE_TTL_SECONDS = 900  # 15 min — live market data stays fresh, falls back gracefully after

# In-memory fallback cache (used when Mongo is unavailable)
_MEMORY_CACHE: dict[str, tuple[datetime, Any]] = {}

# Git SHA for version endpoint
def _get_version() -> str:
    try:
        sha = subprocess.check_output(
            ["git", "rev-parse", "HEAD"],
            cwd=ROOT_DIR,
            stderr=subprocess.DEVNULL,
        ).decode().strip()[:7]
        return sha
    except Exception:
        return os.environ.get("APP_VERSION", "unknown")

VERSION = _get_version()
STARTUP_TIME = datetime.now(timezone.utc).isoformat()

# =====================================================================
# MONGODB — lazy, safe, optional
# =====================================================================

_mongo_client = None
_db = None

def _get_db():
    """Return Mongo db handle, or None if unavailable. Never raises."""
    global _mongo_client, _db
    if _db is not None:
        return _db
    if not MOTOR_AVAILABLE or not MONGO_URL:
        return None
    try:
        _mongo_client = AsyncIOMotorClient(MONGO_URL, serverSelectionTimeoutMS=2000)
        _db = _mongo_client[DB_NAME]
        return _db
    except Exception as e:
        logger.warning(f"Mongo connection failed: {e}")
        return None

# =====================================================================
# CACHE HELPERS — Mongo first, memory fallback
# =====================================================================

async def cache_set(key: str, value: Any) -> None:
    """Store value in cache. Tries Mongo, falls back to memory. Never raises."""
    now = datetime.now(timezone.utc)
    # Always write to memory as safety net
    _MEMORY_CACHE[key] = (now, value)
    # Try Mongo
    db = _get_db()
    if db is None:
        return
    try:
        await db.cache.update_one(
            {"_id": key},
            {"$set": {"value": value, "updated_at": now}},
            upsert=True,
        )
    except Exception as e:
        logger.warning(f"cache_set({key}) mongo failed: {e}")

async def cache_get(key: str, max_age_seconds: int = CACHE_TTL_SECONDS) -> tuple[Optional[Any], bool]:
    """
    Fetch cached value. Returns (value, is_stale).
    If value is None, nothing was cached.
    is_stale=True means the cache is older than max_age_seconds but still usable.
    Never raises.
    """
    # Try Mongo first
    db = _get_db()
    if db is not None:
        try:
            doc = await db.cache.find_one({"_id": key})
            if doc:
                updated = doc.get("updated_at")
                if isinstance(updated, datetime):
                    age = (datetime.now(timezone.utc) - updated.replace(tzinfo=timezone.utc) if updated.tzinfo is None else datetime.now(timezone.utc) - updated).total_seconds()
                    is_stale = age > max_age_seconds
                    return doc.get("value"), is_stale
        except Exception as e:
            logger.warning(f"cache_get({key}) mongo failed: {e}")
    # Memory fallback
    if key in _MEMORY_CACHE:
        cached_at, value = _MEMORY_CACHE[key]
        age = (datetime.now(timezone.utc) - cached_at).total_seconds()
        return value, age > max_age_seconds
    return None, False

# =====================================================================
# SAFE DATA FETCHERS — each one has fallback chain
# =====================================================================

# ---- NSE top movers ----

NIFTY50_TOP = [
    ("RELIANCE", "RELIANCE.NS"),
    ("HDFCBANK", "HDFCBANK.NS"),
    ("ICICIBANK", "ICICIBANK.NS"),
    ("INFY", "INFY.NS"),
    ("TCS", "TCS.NS"),
    ("SBIN", "SBIN.NS"),
    ("BHARTIARTL", "BHARTIARTL.NS"),
    ("HINDUNILVR", "HINDUNILVR.NS"),
    ("BAJFINANCE", "BAJFINANCE.NS"),
    ("KOTAKBANK", "KOTAKBANK.NS"),
]

def _safe_float(x, default=None):
    try:
        if x is None:
            return default
        val = float(x)
        if val != val:  # NaN check
            return default
        return val
    except (TypeError, ValueError):
        return default

async def fetch_nse_movers_safe() -> dict:
    """
    Fetch NSE top movers. Returns:
      {
        "movers": [...],
        "source": "yfinance" | "twelve_data" | "cache" | "skeleton",
        "stale": bool,
      }
    Never raises.
    """
    # Attempt 1: yfinance
    if YF_AVAILABLE:
        try:
            movers = []
            tickers_str = " ".join(t[1] for t in NIFTY50_TOP)
            tks = yf.Tickers(tickers_str)
            for symbol, ticker in NIFTY50_TOP:
                try:
                    tk = tks.tickers.get(ticker) or yf.Ticker(ticker)
                    hist = tk.history(period="2d")
                    price = None
                    prev_close = None
                    volume = None
                    if hist is not None and not hist.empty:
                        price = _safe_float(hist["Close"].iloc[-1])
                        if len(hist) >= 2:
                            prev_close = _safe_float(hist["Close"].iloc[-2])
                        volume = _safe_float(hist["Volume"].iloc[-1])
                    # If history didn't give price, try info as last resort
                    if price is None:
                        try:
                            info = tk.info or {}
                            price = _safe_float(info.get("currentPrice") or info.get("regularMarketPrice"))
                            prev_close = prev_close or _safe_float(info.get("previousClose"))
                            volume = volume or _safe_float(info.get("volume"))
                        except Exception:
                            pass
                    change = None
                    change_pct = None
                    if price is not None and prev_close:
                        change = price - prev_close
                        change_pct = (change / prev_close) * 100 if prev_close else None
                    movers.append({
                        "symbol": symbol,
                        "ticker": ticker,
                        "ltp": price,
                        "price": price,
                        "prev_close": prev_close,
                        "change": change,
                        "change_percent": change_pct,
                        "volume": volume,
                        "changeType": "positive" if (change or 0) >= 0 else "negative",
                    })
                except Exception as e:
                    logger.debug(f"yfinance per-ticker {ticker}: {e}")
                    movers.append(_mover_skeleton(symbol, ticker))

            # Check if we got ANY real prices. If every price is None, yfinance is blocked.
            real_prices = sum(1 for m in movers if m["price"] is not None)
            if real_prices >= 3:
                return {"movers": movers, "source": "yfinance", "stale": False}
            logger.warning(f"yfinance returned only {real_prices}/10 real prices — likely rate limited")
        except Exception as e:
            logger.warning(f"yfinance batch failed: {e}")

    # Attempt 2: Twelve Data (if key configured)
    if TWELVE_DATA_KEY and HTTPX_AVAILABLE:
        try:
            movers = await _fetch_twelve_data_movers()
            if movers and any(m["price"] is not None for m in movers):
                return {"movers": movers, "source": "twelve_data", "stale": False}
        except Exception as e:
            logger.warning(f"twelve_data failed: {e}")

    # Attempt 3: cache
    cached, is_stale = await cache_get("nse_movers")
    if cached:
        return {"movers": cached, "source": "cache", "stale": is_stale}

    # Attempt 4: skeleton (never null frontend)
    return {
        "movers": [_mover_skeleton(sym, tk) for sym, tk in NIFTY50_TOP],
        "source": "skeleton",
        "stale": True,
    }

def _mover_skeleton(symbol: str, ticker: str) -> dict:
    return {
        "symbol": symbol,
        "ticker": ticker,
        "ltp": None,
        "price": None,
        "prev_close": None,
        "change": None,
        "change_percent": None,
        "volume": None,
        "changeType": "neutral",
    }

async def _fetch_twelve_data_movers() -> list[dict]:
    """Fetch NSE prices via Twelve Data API. Returns [] on any failure."""
    if not TWELVE_DATA_KEY or not HTTPX_AVAILABLE:
        return []
    movers = []
    async with httpx.AsyncClient(timeout=10.0) as client:
        for symbol, ticker in NIFTY50_TOP:
            try:
                # Twelve Data uses symbol like "RELIANCE:NSE"
                td_sym = f"{symbol}:NSE"
                r = await client.get(
                    "https://api.twelvedata.com/quote",
                    params={"symbol": td_sym, "apikey": TWELVE_DATA_KEY},
                )
                if r.status_code == 200:
                    d = r.json()
                    if d.get("code"):  # Twelve Data errors come with a code
                        continue
                    price = _safe_float(d.get("close"))
                    prev = _safe_float(d.get("previous_close"))
                    change = _safe_float(d.get("change"))
                    change_pct = _safe_float(d.get("percent_change"))
                    movers.append({
                        "symbol": symbol, "ticker": ticker,
                        "ltp": price, "price": price,
                        "prev_close": prev,
                        "change": change, "change_percent": change_pct,
                        "volume": _safe_float(d.get("volume")),
                        "changeType": "positive" if (change or 0) >= 0 else "negative",
                    })
                else:
                    movers.append(_mover_skeleton(symbol, ticker))
            except Exception as e:
                logger.debug(f"twelve_data {symbol}: {e}")
                movers.append(_mover_skeleton(symbol, ticker))
    return movers

# ---- FX ----

async def fetch_fx_safe() -> dict:
    """Fetch FX rates. Frankfurter API is cloud-friendly, rarely blocked."""
    default = {
        "CNYINR": {"rate": 13.6, "change_percent": 0.0},
        "EURINR": {"rate": 109.5, "change_percent": 0.0},
        "GBPINR": {"rate": 125.6, "change_percent": 0.0},
        "JPYINR": {"rate": 0.58, "change_percent": 0.0},
        "SGDINR": {"rate": 72.9, "change_percent": 0.0},
        "USDINR": {"rate": 92.9, "change_percent": 0.0},
    }
    if not HTTPX_AVAILABLE:
        cached, _ = await cache_get("fx")
        return cached or default
    try:
        async with httpx.AsyncClient(timeout=8.0) as client:
            r = await client.get("https://api.frankfurter.app/latest?from=INR")
            if r.status_code == 200:
                d = r.json()
                rates = d.get("rates", {})
                out = {}
                for ccy in ["CNY", "EUR", "GBP", "JPY", "SGD", "USD"]:
                    rate = rates.get(ccy)
                    if rate:
                        out[f"{ccy}INR"] = {"rate": round(1 / rate, 4), "change_percent": 0.0}
                if out:
                    await cache_set("fx", out)
                    return out
    except Exception as e:
        logger.warning(f"fx fetch failed: {e}")
    cached, _ = await cache_get("fx")
    return cached or default

# ---- Commodities ----

async def fetch_commodities_safe() -> list[dict]:
    default = [
        {"name": "Gold", "symbol": "GC=F", "price": 0, "currency": "USD", "unit": "USD/oz", "change_pct": 0},
        {"name": "Silver", "symbol": "SI=F", "price": 0, "currency": "USD", "unit": "USD/oz", "change_pct": 0},
        {"name": "Crude Oil (WTI)", "symbol": "CL=F", "price": 0, "currency": "USD", "unit": "USD/bbl", "change_pct": 0},
        {"name": "Brent Crude", "symbol": "BZ=F", "price": 0, "currency": "USD", "unit": "USD/bbl", "change_pct": 0},
        {"name": "Natural Gas", "symbol": "NG=F", "price": 0, "currency": "USD", "unit": "USD/MMBtu", "change_pct": 0},
    ]
    if not YF_AVAILABLE:
        cached, _ = await cache_get("commodities")
        return cached or default
    try:
        symbols = ["GC=F", "SI=F", "CL=F", "BZ=F", "NG=F"]
        tks = yf.Tickers(" ".join(symbols))
        out = []
        meta = {
            "GC=F": ("Gold", "USD/oz"),
            "SI=F": ("Silver", "USD/oz"),
            "CL=F": ("Crude Oil (WTI)", "USD/bbl"),
            "BZ=F": ("Brent Crude", "USD/bbl"),
            "NG=F": ("Natural Gas", "USD/MMBtu"),
        }
        for sym in symbols:
            try:
                name, unit = meta[sym]
                tk = tks.tickers.get(sym) or yf.Ticker(sym)
                hist = tk.history(period="2d")
                price = 0
                change_pct = 0
                if hist is not None and not hist.empty:
                    price = _safe_float(hist["Close"].iloc[-1], 0)
                    if len(hist) >= 2:
                        prev = _safe_float(hist["Close"].iloc[-2], price)
                        if prev:
                            change_pct = round((price - prev) / prev * 100, 2)
                out.append({
                    "name": name, "symbol": sym,
                    "price": price, "currency": "USD",
                    "unit": unit, "change_pct": change_pct,
                })
            except Exception as e:
                logger.debug(f"commodity {sym}: {e}")
        if out and any(c["price"] for c in out):
            await cache_set("commodities", out)
            return out
    except Exception as e:
        logger.warning(f"commodities failed: {e}")
    cached, _ = await cache_get("commodities")
    return cached or default

# ---- News ----

async def fetch_news_safe() -> list[dict]:
    default = []
    if not HTTPX_AVAILABLE:
        cached, _ = await cache_get("news")
        return cached or default
    feeds = [
        ("Economic Times Markets", "https://economictimes.indiatimes.com/markets/rssfeeds/1977021501.cms"),
        ("Mint Markets", "https://www.livemint.com/rss/markets"),
    ]
    news = []
    try:
        async with httpx.AsyncClient(timeout=8.0, follow_redirects=True) as client:
            for source, url in feeds:
                try:
                    r = await client.get(url)
                    if r.status_code != 200:
                        continue
                    # Cheap RSS parse without feedparser dep
                    text = r.text
                    items = text.split("<item>")[1:6]
                    for item in items:
                        try:
                            title = _between(item, "<title>", "</title>").replace("<![CDATA[", "").replace("]]>", "").strip()
                            link = _between(item, "<link>", "</link>").strip()
                            pub = _between(item, "<pubDate>", "</pubDate>").strip()
                            if title and link:
                                news.append({"title": title, "source": source, "url": link, "published": pub})
                        except Exception:
                            continue
                except Exception as e:
                    logger.debug(f"news feed {source}: {e}")
        if news:
            await cache_set("news", news)
            return news
    except Exception as e:
        logger.warning(f"news failed: {e}")
    cached, _ = await cache_get("news")
    return cached or default

def _between(s: str, start: str, end: str) -> str:
    try:
        a = s.index(start) + len(start)
        b = s.index(end, a)
        return s[a:b]
    except ValueError:
        return ""

# ---- FII/DII ----
# Fetches daily FII/DII net flows from NSE (JSON) with cookie bootstrap.
# Maintains a 14-day rolling history in cache so the chart has real history
# even though the NSE endpoint only returns the latest trading day.
#
# Units: fii_net / dii_net are in ₹ Crore (NSE's native unit).

_NSE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/122.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.nseindia.com/reports/fii-dii",
}

async def _fetch_fii_dii_nse_latest() -> Optional[dict]:
    """Hit NSE fiidiiTradeReact. Returns {date, fii_net, dii_net} in Cr, or None."""
    if not HTTPX_AVAILABLE:
        return None
    try:
        async with httpx.AsyncClient(timeout=8.0, headers=_NSE_HEADERS, follow_redirects=True) as client:
            # Bootstrap cookies
            await client.get("https://www.nseindia.com/reports/fii-dii")
            r = await client.get("https://www.nseindia.com/api/fiidiiTradeReact")
            if r.status_code != 200:
                logger.warning(f"NSE fii/dii HTTP {r.status_code}")
                return None
            rows = r.json()
            if not isinstance(rows, list) or not rows:
                return None
            fii_net = None
            dii_net = None
            trade_date = None
            for row in rows:
                cat = (row.get("category") or "").upper()
                net = row.get("netValue")
                if isinstance(net, str):
                    try:
                        net = float(net.replace(",", ""))
                    except ValueError:
                        net = None
                if net is None:
                    continue
                if "FII" in cat or "FPI" in cat:
                    fii_net = float(net)
                elif "DII" in cat:
                    dii_net = float(net)
                if not trade_date:
                    trade_date = row.get("date")
            if fii_net is None and dii_net is None:
                return None
            # Normalize date to ISO
            iso_date = None
            if trade_date:
                for fmt in ("%d-%b-%Y", "%d %b %Y", "%Y-%m-%d"):
                    try:
                        iso_date = datetime.strptime(trade_date, fmt).date().isoformat()
                        break
                    except ValueError:
                        continue
            if not iso_date:
                iso_date = datetime.now(timezone.utc).date().isoformat()
            return {
                "date": iso_date,
                "fii_net": round(fii_net or 0.0, 2),
                "dii_net": round(dii_net or 0.0, 2),
                "nifty_close": 0,
            }
    except Exception as e:
        logger.warning(f"NSE fii/dii fetch failed: {e}")
        return None

async def fetch_fii_dii_safe() -> list[dict]:
    """Return recent FII/DII flows as list (latest first), in ₹ Crore."""
    # Load rolling history from cache
    history_raw, _ = await cache_get("fii_dii_history", max_age_seconds=86400 * 30)
    history: list[dict] = history_raw if isinstance(history_raw, list) else []

    # Try live fetch; merge if new date
    latest = await _fetch_fii_dii_nse_latest()
    if latest:
        existing_dates = {row.get("date") for row in history}
        if latest["date"] not in existing_dates:
            history = [latest] + history
        else:
            # update in place (values may have been revised)
            history = [latest if row.get("date") == latest["date"] else row for row in history]
        history = sorted(history, key=lambda r: r.get("date", ""), reverse=True)[:14]
        await cache_set("fii_dii_history", history)

    if history:
        return history

    # No live data and no cache — return zero-filled skeleton so chart renders empty state
    today = datetime.now(timezone.utc).date()
    return [
        {"date": (today - timedelta(days=i)).isoformat(), "fii_net": 0, "dii_net": 0, "nifty_close": 0}
        for i in range(1, 8)
    ]

# =====================================================================
# ENDPOINT ERROR WRAPPER — never let anything 500
# =====================================================================

def safe_endpoint(fallback_factory):
    """
    Decorator: catches any exception from an endpoint, logs it with full
    traceback, and returns a structured 200 with fallback data instead of 500.
    fallback_factory: callable that returns the fallback response dict.

    Uses functools.wraps to preserve the wrapped function's signature so
    FastAPI's dependency injection reads the real parameters, not *args/**kwargs.
    """
    def decorator(fn):
        @functools.wraps(fn)
        async def wrapper(*args, **kwargs):
            try:
                return await fn(*args, **kwargs)
            except HTTPException:
                raise  # let intentional HTTP errors through (e.g., 404 for unknown ticker)
            except Exception as e:
                tb = traceback.format_exc()
                logger.error(f"endpoint {fn.__name__} failed: {e}\n{tb}")
                fallback = fallback_factory() if callable(fallback_factory) else fallback_factory
                return JSONResponse(
                    status_code=200,
                    content={
                        **fallback,
                        "_error": {
                            "message": str(e),
                            "endpoint": fn.__name__,
                            "stage": "handler",
                        },
                        "_fallback": True,
                    },
                )
        return wrapper
    return decorator

# =====================================================================
# APP + ROUTER
# =====================================================================

from contextlib import asynccontextmanager

@asynccontextmanager
async def _lifespan(app):
    # Start APScheduler in background thread on startup
    _sched = None
    try:
        import sys as _sys
        _sys.path.insert(0, str(Path(__file__).parent / "research_platform"))
        from scheduler.runner import build_scheduler as _build_scheduler
        _sched = _build_scheduler()
        _sched.start()
        logger.info(f"Scheduler started with {len(_sched.get_jobs())} jobs")
    except Exception as _se:
        logger.warning(f"Scheduler start failed (non-fatal): {_se}")
    yield
    if _sched is not None:
        try:
            _sched.shutdown(wait=False)
        except Exception:
            pass

app = FastAPI(title="Market Pulse — Hardened Backend", version=VERSION, lifespan=_lifespan)

# CORS as early as possible
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS if CORS_ORIGINS != [""] else ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

api_router = APIRouter(prefix="/api")

# ---------- Meta ----------

@api_router.get("/")
async def root():
    return {"message": "Hello World", "service": "market_pulse", "version": VERSION}

@api_router.get("/health")
async def health():
    db = _get_db()
    # Check DCF notebook prerequisites
    dcf_notebook_present = DCF_NOTEBOOK_SRC.exists() if 'DCF_NOTEBOOK_SRC' in globals() else False
    try:
        import papermill as _pm
        papermill_available = True
    except Exception:
        papermill_available = False
    return {
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "version": VERSION,
        "mongo": "connected" if db is not None else "unavailable",
        "yfinance": "available" if YF_AVAILABLE else "unavailable",
        "twelve_data": "configured" if TWELVE_DATA_KEY else "not_configured",
        "research_platform": {
            "available": RP_AVAILABLE,
            "session_manager": RP_SESSION_MGR,
            "scenarios": RP_SCENARIOS,
            "scoring": RP_SCORING,
            "pdf_builder": RP_PDF,
            "audit_export": RP_EXPORT,
            "errors": _rp_errors if _rp_errors else None,
        },
        "dcf_notebook": {
            "notebook_present": dcf_notebook_present,
            "papermill_available": papermill_available,
            "ready": dcf_notebook_present and papermill_available,
        },
        "uptime_since": STARTUP_TIME,
    }

@api_router.get("/version")
async def version():
    return {"version": VERSION, "startup": STARTUP_TIME}

@api_router.get("/ping")
async def ping():
    return {"ok": True}

# ---------- Market Overview ----------

def _overview_fallback():
    return {
        "top_movers": [_mover_skeleton(s, t) for s, t in NIFTY50_TOP],
        "fx": {},
        "crypto": [],
        "commodities": [],
        "fii_dii": [],
        "news": [],
        "last_updated": datetime.now(timezone.utc).isoformat(),
    }

@api_router.get("/market/overview")
@safe_endpoint(_overview_fallback)
async def market_overview():
    # Fan out all fetches in parallel; each is individually safe
    movers_task = fetch_nse_movers_safe()
    fx_task = fetch_fx_safe()
    commodities_task = fetch_commodities_safe()
    news_task = fetch_news_safe()
    fii_task = fetch_fii_dii_safe()

    movers_res, fx_res, comm_res, news_res, fii_res = await asyncio.gather(
        movers_task, fx_task, commodities_task, news_task, fii_task,
        return_exceptions=True,
    )

    # Unpack with defensive checks
    if isinstance(movers_res, dict):
        top_movers = movers_res.get("movers", [])
        movers_source = movers_res.get("source", "unknown")
        movers_stale = movers_res.get("stale", False)
    else:
        top_movers = [_mover_skeleton(s, t) for s, t in NIFTY50_TOP]
        movers_source = "error"
        movers_stale = True

    fx = fx_res if isinstance(fx_res, dict) else {}
    commodities = comm_res if isinstance(comm_res, list) else []
    news = news_res if isinstance(news_res, list) else []
    fii_dii = fii_res if isinstance(fii_res, list) else []

    indices = await fetch_indices_safe()
    nifty = indices.get("NIFTY50") or indices.get("NIFTY 50") or {}
    sensex = indices.get("SENSEX") or {}

    return {
        "top_movers": top_movers,
        "fx": fx,
        "crypto": [],  # TODO: wire CoinGecko fallback
        "commodities": commodities,
        "fii_dii": fii_dii,
        "news": news,
        "indices": indices,
        # Flat index fields — frontend may look for these directly
        "nifty": nifty,
        "nifty50": nifty,
        "NIFTY50": nifty,
        "sensex": sensex,
        "SENSEX": sensex,
        "nifty_value": nifty.get("value"),
        "nifty_change_percent": nifty.get("change_percent"),
        "sensex_value": sensex.get("value"),
        "sensex_change_percent": sensex.get("change_percent"),
        "last_updated": datetime.now(timezone.utc).isoformat(),
        "_meta": {
            "movers_source": movers_source,
            "movers_stale": movers_stale,
        },
    }

# ---------- Indices (NIFTY 50 + SENSEX) ----------

INDICES = [
    ("NIFTY 50", "^NSEI"),
    ("SENSEX", "^BSESN"),
]

async def fetch_indices_safe() -> dict:
    """Fetch NIFTY 50 and SENSEX levels. Never raises."""
    out = {}
    if YF_AVAILABLE:
        try:
            loop = asyncio.get_event_loop()
            def _fetch_one(name: str, symbol: str):
                try:
                    tk = yf.Ticker(symbol)
                    hist = tk.history(period="2d")
                    if hist is not None and not hist.empty:
                        price = _safe_float(hist["Close"].iloc[-1])
                        prev = _safe_float(hist["Close"].iloc[-2]) if len(hist) >= 2 else None
                        change = (price - prev) if (price and prev) else None
                        change_pct = ((change / prev) * 100) if (change and prev) else None
                        change_type = "neutral"
                        if change is not None:
                            change_type = "positive" if change >= 0 else "negative"
                        # Pre-formatted display strings for the pill
                        value_fmt = f"{price:,.2f}" if price is not None else "—"
                        change_fmt = f"{change_pct:+.2f}%" if change_pct is not None else "—"
                        return name, {
                            "name": name,
                            "symbol": symbol,
                            "value": value_fmt,           # display string "24,353.55"
                            "raw_value": price,           # raw number for toLocaleString
                            "price": price,               # alias
                            "ltp": price,                 # alias
                            "last": price,                # alias
                            "prev_close": prev,
                            "change": change_fmt,         # display string "+0.65%" — frontend renders this
                            "change_raw": change,         # raw delta number
                            "change_percent": change_pct,
                            "change_pct": change_pct,     # alias
                            "changeType": change_type,    # "positive" / "negative" / "neutral"
                        }
                except Exception as e:
                    logger.debug(f"indices {symbol}: {e}")
                return name, {"name": name, "symbol": symbol, "value": "—",
                              "raw_value": None, "price": None, "ltp": None, "last": None,
                              "prev_close": None, "change": "—", "change_raw": None,
                              "change_percent": None, "change_pct": None,
                              "changeType": "neutral"}

            results = await asyncio.gather(
                *[loop.run_in_executor(None, _fetch_one, n, s) for n, s in INDICES],
                return_exceptions=True,
            )
            for res in results:
                if isinstance(res, tuple):
                    name, data = res
                    # Expose under every common key the frontend might read:
                    # "NIFTY 50" (with space), "NIFTY50" (no space), "nifty" (lowercase)
                    key_nospace = name.replace(" ", "")
                    out[name] = data                        # "NIFTY 50" / "SENSEX"
                    out[key_nospace] = data                 # "NIFTY50" / "SENSEX"
                    out[key_nospace.lower()] = data         # "nifty50" / "sensex"
                    # Special shortname for NIFTY 50 → "nifty" (TopBar.jsx uses this)
                    if "NIFTY" in key_nospace:
                        out["nifty"] = data
                    if "SENSEX" in key_nospace:
                        out["sensex"] = data
            if out:
                await cache_set("indices", out)
                return out
        except Exception as e:
            logger.warning(f"indices fetch failed: {e}")
    cached, _ = await cache_get("indices")
    return cached or {
        "NIFTY50": {"name": "NIFTY 50", "value": None, "change_percent": None},
        "nifty": {"name": "NIFTY 50", "value": None, "change_percent": None},
        "SENSEX": {"name": "SENSEX", "value": None, "change_percent": None},
        "sensex": {"name": "SENSEX", "value": None, "change_percent": None},
    }

# ---------- Sessions ----------

class SessionCreate(BaseModel):
    ticker: str
    sector: Optional[str] = None

@api_router.get("/sessions")
@safe_endpoint(lambda: {"sessions": []})
async def list_sessions():
    """List research sessions, deduplicated by ticker (keep newest)."""
    db = _get_db()
    if db is None:
        return {"sessions": []}
    try:
        # Dedupe pipeline: group by ticker, keep newest
        pipeline = [
            {"$sort": {"created_at": -1}},
            {"$group": {
                "_id": "$ticker",
                "doc": {"$first": "$$ROOT"},
            }},
            {"$replaceRoot": {"newRoot": "$doc"}},
            {"$sort": {"created_at": -1}},
            {"$limit": 50},
        ]
        cursor = db.sessions.aggregate(pipeline)
        docs = []
        async for d in cursor:
            d["_id"] = str(d.get("_id", ""))
            docs.append(d)
        return {"sessions": docs}
    except Exception as e:
        logger.warning(f"list_sessions mongo: {e}")
        return {"sessions": []}

@api_router.post("/research/new")
@safe_endpoint(lambda: {"session_id": "offline", "ticker": "UNKNOWN", "created_at": datetime.now(timezone.utc).isoformat()})
async def create_session(req: SessionCreate):
    ticker = (req.ticker or "").upper().strip()
    if not ticker:
        raise HTTPException(status_code=400, detail="ticker required")
    now = datetime.now(timezone.utc)
    session_id = now.strftime("%y%m%d_%H%M%S")
    doc = {
        "_id": session_id,
        "session_id": session_id,
        "ticker": ticker,
        "sector": req.sector or "general",
        "hypothesis": f"Analysis session for {ticker}",
        "status": "active",
        "created_at": now,
    }
    db = _get_db()
    if db is not None:
        try:
            await db.sessions.insert_one(doc)
        except Exception as e:
            logger.warning(f"create_session insert: {e}")
    doc["created_at"] = now.isoformat()
    return doc

# ---------- Research Analyze (the endpoint that was 500-ing) ----------

def _analyze_fallback():
    return {
        "ticker": "UNKNOWN",
        "price_data": None,
        "dcf": None,
        "scenarios": [],
        "message": "Analysis unavailable — using cached or fallback data",
    }

@api_router.post("/research/analyze")
@safe_endpoint(_analyze_fallback)
async def analyze(request: Request):
    """
    The previously-500ing endpoint. Now:
      - Every numerical operation guarded with safe_float() + default
      - Missing yfinance data → skeleton response, not exception
      - Division by zero → guarded everywhere
      - Detailed error surface via safe_endpoint decorator
    """
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    ticker = (body.get("ticker") or "").upper().strip()
    if not ticker:
        raise HTTPException(status_code=400, detail="ticker required")

    ticker_ns = ticker if "." in ticker else f"{ticker}.NS"

    # Try cache first for instant response
    cache_key = f"analyze:{ticker_ns}"

    # Pull live data defensively
    price_data = None
    info = {}
    if YF_AVAILABLE:
        try:
            tk = yf.Ticker(ticker_ns)
            hist = tk.history(period="2d")
            if hist is not None and not hist.empty:
                price = _safe_float(hist["Close"].iloc[-1])
                prev = _safe_float(hist["Close"].iloc[-2]) if len(hist) >= 2 else None
                price_data = {
                    "price": price,
                    "prev_close": prev,
                    "change": (price - prev) if (price and prev) else None,
                    "change_pct": ((price - prev) / prev * 100) if (price and prev) else None,
                }
            try:
                info = tk.info or {}
            except Exception:
                info = {}
        except Exception as e:
            logger.warning(f"analyze yfinance failed for {ticker}: {e}")

    # Don't early-return from cache — we always want to go through the full
    # response-building path below so all fields (session_id, flat prices,
    # etc.) are present. Cache is only used when yfinance fully crashes.

    # If history() returned null price but info has currentPrice, use that
    if price_data is None or price_data.get("price") is None:
        info_price = _safe_float(info.get("currentPrice") or info.get("regularMarketPrice"))
        info_prev = _safe_float(info.get("previousClose") or (price_data or {}).get("prev_close"))
        if info_price is not None:
            price_data = {
                "price": info_price,
                "prev_close": info_prev,
                "change": (info_price - info_prev) if (info_prev is not None) else None,
                "change_pct": ((info_price - info_prev) / info_prev * 100) if (info_prev) else None,
            }

    # Safe DCF skeleton — never divides by zero
    current_price = _safe_float(info.get("currentPrice") or (price_data or {}).get("price"), 0)
    market_cap = _safe_float(info.get("marketCap"), 0)
    total_revenue = _safe_float(info.get("totalRevenue"), 0)
    net_income = _safe_float(info.get("netIncomeToCommon") or info.get("netIncome"), 0)
    shares_out = _safe_float(info.get("sharesOutstanding"), 0)
    beta = _safe_float(info.get("beta"), 1.0)

    # Safe WACC (all bounds guarded)
    rf = 0.072  # India 10Y
    erp = 0.06
    wacc = min(max(rf + beta * erp, 0.08), 0.15)

    # Safe terminal value
    growth = 0.05
    terminal_growth = 0.035
    if wacc <= terminal_growth:
        wacc = terminal_growth + 0.02  # prevent division issues

    # Simple 5-year FCF projection (all zero-safe)
    fcf_base = net_income * 0.7 if net_income else 0
    pv_fcf = 0
    for year in range(1, 6):
        fcf_year = fcf_base * ((1 + growth) ** year)
        pv_fcf += fcf_year / ((1 + wacc) ** year)
    terminal_value = 0
    if fcf_base and (wacc - terminal_growth) > 0:
        terminal_fcf = fcf_base * ((1 + growth) ** 5) * (1 + terminal_growth)
        terminal_value = (terminal_fcf / (wacc - terminal_growth)) / ((1 + wacc) ** 5)
    enterprise_value = pv_fcf + terminal_value

    # Equity value = EV (we skip debt adj in backup for safety)
    equity_value = enterprise_value
    price_per_share = (equity_value / shares_out) if shares_out else 0
    upside_pct = ((price_per_share - current_price) / current_price * 100) if current_price else 0

    # ═══════════════════════════════════════════════════════════════════
    # RESEARCH PLATFORM integration — the real engine, guarded by try/except
    # ═══════════════════════════════════════════════════════════════════
    rp_session_id = None
    rp_scenarios = None
    rp_scoring_data = None
    rp_assumptions = None
    rp_sector = _detect_sector_simple(ticker)
    rp_signals: list = []
    rp_factor_deltas: list = []
    rp_swot_data = None
    rp_porter_data = None
    rp_audit_entries: list = []

    if RP_AVAILABLE:
        try:
            # Create a real research session folder on disk
            rp_ses = rp_new_session(
                ticker=ticker,
                hypothesis=body.get("hypothesis") or f"Analysis session for {ticker}",
                variant_view=body.get("variant_view") or "",
                catalysts=body.get("catalysts") or [],
            )
            rp_session_id = rp_ses.session_id

            # Build assumptions WITHOUT the DB-dependent dcf_bridge.
            # Use actuals from yfinance + sector defaults.
            rp_assumptions = _build_rp_assumptions(
                ticker=ticker,
                sector=rp_sector,
                current_price=current_price,
                market_cap=market_cap,
                total_revenue=total_revenue,
                net_income=net_income,
                shares_out=shares_out,
                beta=beta,
                info=info,
            )
            rp_ses.initialize_assumptions(rp_assumptions)

            # Run scenarios (Bull / Base / Bear + sensitivity + reverse DCF)
            if RP_SCENARIOS and rp_run_scenarios:
                try:
                    base_revenue_cr = (total_revenue / 1e7) if total_revenue else 100.0
                    scenario_result = rp_run_scenarios(
                        rp_ses,
                        rp_assumptions,
                        shares_outstanding=(shares_out / 1e7) if shares_out else None,
                        base_revenue=max(base_revenue_cr, 1.0),
                    )
                    # Read back the structured JSON that was written to disk
                    rp_scenarios = rp_ses.get_scenarios()
                    logger.info(f"[analyze] RP scenarios ran for {ticker}")
                except Exception as e:
                    logger.warning(f"[analyze] RP scenarios failed: {e}")

            # Run scoring (5-dimension composite)
            scoring = None
            if RP_SCORING and rp_score_session:
                try:
                    scoring = rp_score_session(rp_ses, sector=rp_sector)
                    rp_scoring_data = {
                        "composite_score": round(scoring.composite_score, 1),
                        "recommendation": scoring.recommendation,
                        "business_quality": scoring.business_quality,
                        "financial_strength": round(scoring.financial_strength, 1),
                        "growth_quality": round(scoring.growth_quality, 1),
                        "valuation_attractiveness": round(scoring.valuation_attractiveness, 1),
                        "risk_score": round(scoring.risk_score, 1),
                        "market_positioning": round(scoring.market_positioning, 1),
                        "rationale": scoring.rationale[:5],
                        "caveats": scoring.caveats[:5],
                    }
                    logger.info(f"[analyze] RP scoring: {scoring.recommendation} ({scoring.composite_score:.0f})")
                except Exception as e:
                    logger.warning(f"[analyze] RP scoring failed: {e}")

            # ── 1. Detect signals from sector context + news headlines ──
            if RP_SIGNALS and rp_detect_signals:
                try:
                    _sector_text = f"{ticker} {rp_sector} sector stock analysis"
                    rp_signals = rp_detect_signals(
                        _sector_text, "sector_context",
                        ticker=ticker, sector=rp_sector,
                    ) or []
                    try:
                        _news = await fetch_news_safe()
                        for _item in (_news or [])[:10]:
                            _title = _item.get("title", "")
                            if _title and ticker.lower() in _title.lower():
                                _sigs = rp_detect_signals(
                                    _title, _item.get("source", "news"),
                                    ticker=ticker, sector=rp_sector,
                                )
                                rp_signals.extend(_sigs or [])
                    except Exception as _ne:
                        logger.debug(f"[analyze] news signals skipped: {_ne}")
                    logger.info(f"[analyze] {ticker}: {len(rp_signals)} signals detected")
                except Exception as _e:
                    logger.warning(f"[analyze] detect_signals failed: {_e}")

            # ── 2. Map signals → factor assumption deltas ──
            if RP_FACTORS and rp_signals_to_factors:
                try:
                    _cur_asmp: dict = {}
                    if rp_scenarios and isinstance(rp_scenarios.get("base"), dict):
                        _cur_asmp = rp_scenarios["base"].get("assumptions", {}) or {}
                    elif rp_assumptions:
                        _cur_asmp = {k: v for k, v in rp_assumptions.items()
                                     if isinstance(v, (int, float))}
                    rp_factor_deltas = rp_signals_to_factors(rp_signals, _cur_asmp) or []
                    logger.info(f"[analyze] {ticker}: {len(rp_factor_deltas)} factor deltas")
                except Exception as _e:
                    logger.warning(f"[analyze] signals_to_factors failed: {_e}")

            # ── 3. SWOT (needs real scoring object) ──
            if RP_SWOT and rp_generate_swot and scoring is not None:
                try:
                    rp_swot_data = rp_generate_swot(rp_ses, scoring, sector=rp_sector)
                    logger.info(f"[analyze] {ticker}: SWOT generated")
                except Exception as _e:
                    logger.warning(f"[analyze] generate_swot failed: {_e}")

            # ── 4. Porter Five Forces ──
            if RP_PORTER and rp_generate_porter:
                try:
                    rp_porter_data = rp_generate_porter(rp_ses, sector=rp_sector)
                    logger.info(f"[analyze] {ticker}: Porter generated")
                except Exception as _e:
                    logger.warning(f"[analyze] generate_porter failed: {_e}")

            # ── 5. Write audit events + persist audit_log.json ──
            try:
                rp_ses.audit("session_created", f"Research session opened for {ticker}")
                if rp_scoring_data:
                    rp_ses.audit("scoring_complete",
                        f"Composite score {rp_scoring_data.get('composite_score')} — "
                        f"{rp_scoring_data.get('recommendation')}")
                if rp_signals:
                    rp_ses.audit("signals_detected", f"{len(rp_signals)} signals detected")
                if rp_factor_deltas:
                    rp_ses.audit("factors_mapped", f"{len(rp_factor_deltas)} factor deltas")
                if rp_swot_data:
                    rp_ses.audit("swot_generated", "SWOT analysis complete")
                if rp_porter_data:
                    rp_ses.audit("porter_generated", "Porter Five Forces analysis complete")
                rp_ses.audit("pipeline_complete", f"Full analyze pipeline finished for {ticker}")
                from ai_engine.audit_export import get_full_audit as _get_full_audit
                rp_audit_entries = _get_full_audit(rp_ses) or []
                logger.info(f"[analyze] {ticker}: audit {len(rp_audit_entries)} entries persisted")
            except Exception as _e:
                logger.warning(f"[analyze] audit persist failed: {_e}")

            # ── 6. Build and persist summary.md ──
            try:
                _swot = rp_swot_data or {}
                _porter = rp_porter_data or {}
                _sc = rp_scoring_data or {}
                _scenarios = (rp_scenarios or {}).get("scenarios") or {}
                _hypothesis = body.get("hypothesis") or f"Analysis session for {ticker}"

                def _swot_rows(items):
                    return "\n".join(f"- {i}" for i in (items or [])) or "- N/A"

                _md_lines = [
                    f"# {ticker} — Research Summary",
                    f"",
                    f"**Session:** {rp_ses.session_id}  ",
                    f"**Sector:** {rp_sector}  ",
                    f"**Recommendation:** {_sc.get('recommendation', 'N/A')}  ",
                    f"**Composite Score:** {_sc.get('composite_score', 'N/A')}  ",
                    f"",
                    f"## Investment Thesis",
                    f"",
                    f"{_hypothesis}",
                    f"",
                    f"## Scoring",
                    f"",
                    f"| Dimension | Score |",
                    f"|---|---|",
                    f"| Business Quality | {_sc.get('business_quality', 'N/A')} |",
                    f"| Financial Strength | {_sc.get('financial_strength', 'N/A')} |",
                    f"| Growth Quality | {_sc.get('growth_quality', 'N/A')} |",
                    f"| Valuation Attractiveness | {_sc.get('valuation_attractiveness', 'N/A')} |",
                    f"| Risk Score | {_sc.get('risk_score', 'N/A')} |",
                    f"| Market Positioning | {_sc.get('market_positioning', 'N/A')} |",
                    f"",
                ]
                if _swot:
                    _md_lines += [
                        f"## SWOT Analysis",
                        f"",
                        f"### Strengths",
                        _swot_rows(_swot.get("strengths")),
                        f"",
                        f"### Weaknesses",
                        _swot_rows(_swot.get("weaknesses")),
                        f"",
                        f"### Opportunities",
                        _swot_rows(_swot.get("opportunities")),
                        f"",
                        f"### Threats",
                        _swot_rows(_swot.get("threats")),
                        f"",
                    ]
                if _porter:
                    _md_lines += [
                        f"## Porter's Five Forces",
                        f"",
                        f"| Force | Rating | Comment |",
                        f"|---|---|---|",
                    ]
                    for _force, _val in _porter.items():
                        if isinstance(_val, dict):
                            _md_lines.append(
                                f"| {_force.replace('_', ' ').title()} "
                                f"| {_val.get('rating', '')} "
                                f"| {_val.get('comment', '')} |"
                            )
                    _md_lines.append("")
                if rp_signals:
                    _md_lines += [
                        f"## Signals ({len(rp_signals)})",
                        f"",
                    ]
                    for _sig in rp_signals[:10]:
                        _sig_d = _sig if isinstance(_sig, dict) else getattr(_sig, '__dict__', {})
                        _md_lines.append(
                            f"- **{_sig_d.get('signal_type', '')}** "
                            f"({_sig_d.get('direction', '')}): "
                            f"{_sig_d.get('description', _sig_d.get('detail', ''))}"
                        )
                    _md_lines.append("")
                if _scenarios:
                    _md_lines += [f"## Scenarios", f""]
                    for _label in ("bull", "base", "bear"):
                        _sc_item = _scenarios.get(_label)
                        if _sc_item:
                            _md_lines.append(
                                f"- **{_label.title()}**: "
                                f"Fair value ₹{_sc_item.get('fair_value', 'N/A')} "
                                f"(upside {_sc_item.get('upside_pct', 'N/A')}%)"
                            )
                    _md_lines.append("")
                _md_lines += [
                    f"## Rationale",
                    f"",
                ]
                for _r in (_sc.get("rationale") or []):
                    _md_lines.append(f"- {_r}")
                _md_lines.append("")
                rp_ses.write_summary("\n".join(_md_lines))
                logger.info(f"[analyze] {ticker}: summary.md written")
            except Exception as _e:
                logger.warning(f"[analyze] write_summary failed: {_e}")

        except Exception as e:
            logger.error(f"[analyze] RP integration failed, falling back: {e}")
            rp_session_id = None

    # Create Mongo session (use RP session_id if we got one, else fallback)
    now = datetime.now(timezone.utc)
    session_id = rp_session_id or now.strftime("%y%m%d_%H%M%S")
    session_doc = {
        "_id": session_id,
        "session_id": session_id,
        "ticker": ticker,
        "sector": rp_sector,
        "hypothesis": body.get("hypothesis") or f"Analysis session for {ticker}",
        "variant_view": body.get("variant_view") or "",
        "status": "active",
        "created_at": now,
        "rp_session_id": rp_session_id,  # link to filesystem session
    }
    db = _get_db()
    if db is not None:
        try:
            await db.sessions.insert_one(session_doc)
        except Exception as e:
            logger.debug(f"analyze session insert: {e}")

    response = {
        "session_id": session_id,
        "id": session_id,  # some frontends look for .id instead of .session_id
        "ticker": ticker,
        "ticker_ns": ticker_ns,
        "sector": rp_sector,
        "hypothesis": body.get("hypothesis") or f"Analysis session for {ticker}",
        "variant_view": body.get("variant_view") or "",
        "status": "active",
        "created_at": now.isoformat(),
        "rp_session_id": rp_session_id,
        # Flat price fields — frontend may look for any of these
        "price": current_price,
        "ltp": current_price,
        "last_price": current_price,
        "current_price": current_price,
        "prev_close": (price_data or {}).get("prev_close"),
        "change": (price_data or {}).get("change"),
        "change_percent": (price_data or {}).get("change_pct"),
        "change_pct": (price_data or {}).get("change_pct"),
        # Also keep nested shape for consumers that expect it
        "price_data": price_data,
        # Use RP scenarios if available, else fall back to simple DCF
        "scenarios": (rp_scenarios.get("scenarios") if rp_scenarios else None),
        "sensitivity": (rp_scenarios.get("sensitivity") if rp_scenarios else None),
        "reverse_dcf": (rp_scenarios.get("reverse_dcf") if rp_scenarios else None),
        "scoring": rp_scoring_data,
        "factor_scores": (
            {
                "momentum": round(min(100, max(0, rp_scoring_data["growth_quality"] * 0.6 + rp_scoring_data["market_positioning"] * 0.4)), 1),
                "value":    round(min(100, max(0, rp_scoring_data["valuation_attractiveness"])), 1),
                "quality":  round(min(100, max(0, rp_scoring_data["financial_strength"] * 0.6 + rp_scoring_data["market_positioning"] * 0.4)), 1),
                "macro":    round(min(100, max(0, 100 - rp_scoring_data["risk_score"])), 1),
            }
            if rp_scoring_data else None
        ),
        "signals": [dataclasses.asdict(s) for s in rp_signals],
        "factor_deltas": [dataclasses.asdict(d) for d in rp_factor_deltas],
        "swot": rp_swot_data,
        "porter": rp_porter_data,
        "audit_trail": rp_audit_entries,
        "assumption_confidence": (rp_assumptions or {}).get("_confidence_tags") or {},
        "dcf": {
            "current_price": current_price,
            "fair_value": round(price_per_share, 2),
            "upside_pct": round(upside_pct, 2),
            "wacc": round(wacc, 4),
            "terminal_growth": terminal_growth,
            "growth": growth,
            "enterprise_value": round(enterprise_value, 0),
            "equity_value": round(equity_value, 0),
        },
        "meta": {
            "data_source": "yfinance" if info else "skeleton",
            "fallback_used": not bool(info),
            "research_platform": bool(rp_session_id),
            "has_scenarios": bool(rp_scenarios),
            "has_scoring": bool(rp_scoring_data),
            "has_signals": len(rp_signals) > 0,
            "has_factor_deltas": len(rp_factor_deltas) > 0,
            "has_swot": bool(rp_swot_data),
            "has_porter": bool(rp_porter_data),
        },
    }

    # ── Run DCF notebook synchronously and swap in real fair value ────
    try:
        loop = asyncio.get_event_loop()
        nb_result = await asyncio.wait_for(
            loop.run_in_executor(
                None, _execute_notebook_blocking, session_id, ticker, 30
            ),
            timeout=30,
        )
        logger.info(f"[analyze] notebook status={nb_result.get('status')} elapsed={nb_result.get('elapsed')}s")
    except Exception as _e:
        logger.warning(f"[analyze] DCF notebook run failed/timed out: {_e}")

    # Attach DCF summary JSON (papermill output) if available
    try:
        response["dcf_summary"] = _load_dcf_summary(ticker)
    except Exception as _e:
        logger.debug(f"dcf_summary attach failed: {_e}")
        response["dcf_summary"] = None

    # Wire Valuation card to Excel cell P30 (Implied Share Price, DCF sheet).
    # No fallback — if the Excel is missing or unreadable, surface a 503.
    _fv = _get_dcf_excel_p30(ticker)
    if _fv is None:
        raise HTTPException(status_code=503, detail="DCF Excel not available — try Analyze again")
    _fv = float(_fv)
    _cur = current_price
    _ups = ((_fv - float(_cur)) / float(_cur)) * 100.0 if _cur else None
    if _ups is None:
        _rating = "HOLD"
    elif _ups > 15:
        _rating = "BUY"
    elif _ups > -15:
        _rating = "HOLD"
    else:
        _rating = "SELL"
    _scen = response.get("scenarios") or {}
    _base = dict(_scen.get("base") or {})
    _base["price_per_share"] = round(_fv, 2)
    _base["per_share"] = round(_fv, 2)
    if _ups is not None:
        _base["upside_pct"] = round(_ups, 2)
    _base["rating"] = _rating
    _base["_source"] = "dcf_excel_P30"
    _scen["base"] = _base
    response["scenarios"] = _scen
    if response.get("dcf") is not None:
        response["dcf"]["fair_value"] = round(_fv, 2)
        response["dcf"]["_source"] = "dcf_excel_P30"
        if _ups is not None:
            response["dcf"]["upside_pct"] = round(_ups, 2)
    logger.info(
        f"[analyze] {ticker}: excel_P30={_fv:.2f} "
        f"(upside {_ups if _ups is None else f'{_ups:.1f}%'}, rating {_rating})"
    )

    # Cache for next time yfinance fails
    await cache_set(cache_key, response)
    # Also cache by session_id so GET /research/{session_id} works
    await cache_set(f"session:{session_id}", response)
    return response


# ═══════════════════════════════════════════════════════════════════
# RESEARCH PLATFORM helpers
# ═══════════════════════════════════════════════════════════════════

_SECTOR_TICKER_MAP = {
    "petroleum_energy": ["RELIANCE", "BPCL", "IOCL", "IOC", "HPCL", "ONGC", "GAIL", "OIL"],
    "banking_nbfc": ["HDFCBANK", "ICICIBANK", "SBIN", "AXISBANK", "KOTAKBANK",
                     "BAJFINANCE", "BAJAJFINSV", "INDUSINDBK", "IDFCFIRSTB", "FEDERALBNK",
                     "PNB", "BANKBARODA", "IRFC", "RECLTD", "PFC", "HDFCLIFE", "ICICIGI"],
    "it_tech": ["TCS", "INFY", "WIPRO", "HCLTECH", "TECHM", "LTIM", "PERSISTENT", "COFORGE"],
    "pharma": ["SUNPHARMA", "DRREDDY", "CIPLA", "LUPIN", "DIVISLAB", "AUROPHARMA", "TORNTPHARM"],
    "fmcg_retail": ["HINDUNILVR", "ITC", "NESTLEIND", "DABUR", "MARICO", "BRITANNIA", "GODREJCP"],
    "real_estate": ["DLF", "GODREJPROP", "OBEROIRLTY", "PRESTIGE", "BRIGADE", "SOBHA"],
    "auto": ["TATAMOTORS", "MARUTI", "M&M", "BAJAJ-AUTO", "HEROMOTOCO", "EICHERMOT", "TVSMOTOR"],
}


def _detect_sector_simple(ticker: str) -> str:
    """Map ticker to sector. Falls back to 'other' if not in map."""
    t = ticker.upper().replace(".NS", "").replace(".BO", "")
    for sector, tickers in _SECTOR_TICKER_MAP.items():
        if t in tickers:
            return sector
    return "other"


def _build_rp_assumptions(
    ticker: str,
    sector: str,
    current_price: float,
    market_cap: float,
    total_revenue: float,
    net_income: float,
    shares_out: float,
    beta: float,
    info: dict,
) -> dict:
    """
    Build the assumptions dict that scenario_engine.run_scenarios() and
    scoring.score_session() expect. Uses yfinance actuals + sector defaults.
    DOES NOT touch the database (dcf_bridge.pull_live_inputs) since that
    would require a configured DATABASE_URL which Emergent doesn't have.
    """
    # Sector-default WACCs and terminal growth
    sector_wacc = {
        "petroleum_energy": 10.5, "banking_nbfc": 12.0, "pharma": 11.0,
        "it_tech": 11.5, "fmcg_retail": 10.0, "auto": 11.0,
        "real_estate": 12.5, "other": 11.0,
    }.get(sector, 11.0)
    sector_tg = {
        "petroleum_energy": 2.5, "banking_nbfc": 4.0, "pharma": 4.0,
        "it_tech": 4.5, "fmcg_retail": 5.0, "auto": 3.5,
        "real_estate": 3.5, "other": 3.5,
    }.get(sector, 3.5)

    # Derive basic financial metrics from yfinance
    ebitda = _safe_float(info.get("ebitda"), 0)
    ebitda_margin = (ebitda / total_revenue * 100) if total_revenue else None
    if ebitda_margin is None or ebitda_margin <= 0:
        ebitda_margin = {
            "petroleum_energy": 12.0, "banking_nbfc": 30.0, "pharma": 22.0,
            "it_tech": 23.0, "fmcg_retail": 20.0, "auto": 13.0,
            "real_estate": 25.0, "other": 18.0,
        }.get(sector, 18.0)

    revenue_growth = _safe_float(info.get("revenueGrowth"), 0) * 100 if info.get("revenueGrowth") else None
    if revenue_growth is None:
        revenue_growth = 8.0
    revenue_growth = max(-10.0, min(30.0, revenue_growth))

    net_debt = max(_safe_float(info.get("totalDebt"), 0) - _safe_float(info.get("totalCash"), 0), 0)

    return {
        "current_price_inr": current_price,
        "market_cap": market_cap,
        "base_revenue": (total_revenue / 1e7) if total_revenue else 100.0,  # ₹ Cr
        "shares_outstanding": (shares_out / 1e7) if shares_out else None,   # Cr
        "net_debt": (net_debt / 1e7) if net_debt else 0.0,                  # ₹ Cr
        "beta": beta,
        "revenue_growth": revenue_growth,
        "revenue_growth_y1": revenue_growth,
        "revenue_growth_y2": revenue_growth * 0.95,
        "revenue_growth_y3": revenue_growth * 0.90,
        "revenue_growth_y4": revenue_growth * 0.85,
        "revenue_growth_y5": revenue_growth * 0.80,
        "ebitda_margin": ebitda_margin,
        "ebit_margin": ebitda_margin * 0.85,
        "gross_margin": min(ebitda_margin + 12, 70),
        "capex_pct_revenue": 5.0,
        "wacc": sector_wacc,
        "terminal_growth_rate": sector_tg,
        "cost_of_debt": 8.5,
        "equity_risk_premium": 6.5,
        "risk_free_rate": 7.2,
        "tax_rate": 25.0,
        "working_capital_days": 45.0,
        "debt_equity_ratio": 0.3,
        "_sector": sector,
        "_data_source": "yfinance",
        "_confidence_tags": {
            "revenue_growth": "medium",
            "ebitda_margin": "medium",
            "wacc": "medium",
            "terminal_growth_rate": "medium",
        },
    }


# ---------- Other stubs for completeness (never 500) ----------

async def _enrich_with_live_price(session: dict) -> dict:
    """Add live price fields to a session doc so frontend always has data."""
    ticker = session.get("ticker", "").upper()
    if not ticker or ticker == "UNKNOWN":
        return session
    ticker_ns = ticker if "." in ticker else f"{ticker}.NS"
    price = None
    prev_close = None
    if YF_AVAILABLE:
        try:
            loop = asyncio.get_event_loop()
            def _fetch():
                tk = yf.Ticker(ticker_ns)
                hist = tk.history(period="2d")
                p = _safe_float(hist["Close"].iloc[-1]) if (hist is not None and not hist.empty) else None
                pc = _safe_float(hist["Close"].iloc[-2]) if (hist is not None and len(hist) >= 2) else None
                return p, pc
            price, prev_close = await loop.run_in_executor(None, _fetch)
        except Exception as e:
            logger.debug(f"enrich price {ticker}: {e}")
    # Cache fallback
    if price is None:
        cached, _ = await cache_get(f"session:{session.get('session_id') or session.get('_id')}")
        if cached and cached.get("price") is not None:
            return {**session, **{
                "price": cached.get("price"),
                "ltp": cached.get("price"),
                "last_price": cached.get("price"),
                "current_price": cached.get("price"),
                "prev_close": cached.get("prev_close"),
                "change": cached.get("change"),
                "change_percent": cached.get("change_percent"),
                "change_pct": cached.get("change_pct"),
            }}
    change = (price - prev_close) if (price is not None and prev_close) else None
    change_pct = (change / prev_close * 100) if (change is not None and prev_close) else None
    return {**session, **{
        "price": price,
        "ltp": price,
        "last_price": price,
        "current_price": price,
        "prev_close": prev_close,
        "change": change,
        "change_percent": change_pct,
        "change_pct": change_pct,
    }}

# GET session detail — frontend navigates here after analyze
@api_router.get("/research/{session_id}")
@safe_endpoint(lambda: {"session_id": "unknown", "ticker": "UNKNOWN", "status": "not_found", "_fallback": True})
async def get_session(session_id: str):
    # 1. Try Mongo
    db = _get_db()
    doc = None
    if db is not None:
        try:
            doc = await db.sessions.find_one({"_id": session_id})
            if doc:
                doc["_id"] = str(doc.get("_id", ""))
                if isinstance(doc.get("created_at"), datetime):
                    doc["created_at"] = doc["created_at"].isoformat()
        except Exception as e:
            logger.debug(f"get_session mongo: {e}")

    # 2. Merge cached analyze response (has scenarios/reverse_dcf/scoring/sensitivity)
    cached, _ = await cache_get(f"session:{session_id}")
    if doc is None and cached:
        doc = dict(cached)
    elif doc is not None and cached:
        for k, v in cached.items():
            if doc.get(k) in (None, "", [], {}) and v not in (None, "", [], {}):
                doc[k] = v

    if doc is None:
        return {
            "session_id": session_id, "id": session_id,
            "ticker": "UNKNOWN", "sector": "general",
            "status": "not_found", "hypothesis": "Session not found",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }

    # 3. Fall back to RP filesystem session for scenarios/sensitivity/reverse_dcf/scoring
    try:
        if not (doc.get("scenarios") and doc.get("reverse_dcf")):
            rp_ses = _find_rp_session(session_id) or (
                _find_rp_session_by_ticker(doc.get("ticker")) if doc.get("ticker") else None
            )
            if rp_ses is not None:
                try:
                    rp_scen = rp_ses.get_scenarios() if hasattr(rp_ses, "get_scenarios") else None
                except Exception:
                    rp_scen = None
                if isinstance(rp_scen, dict):
                    for src_key, dst_key in [
                        ("scenarios", "scenarios"),
                        ("sensitivity", "sensitivity"),
                        ("reverse_dcf", "reverse_dcf"),
                    ]:
                        if not doc.get(dst_key) and rp_scen.get(src_key):
                            doc[dst_key] = rp_scen[src_key]
                try:
                    rp_sc = rp_ses.get_scoring() if hasattr(rp_ses, "get_scoring") else None
                    if rp_sc and not doc.get("scoring"):
                        doc["scoring"] = rp_sc
                except Exception:
                    pass
    except Exception as e:
        logger.debug(f"get_session RP enrich: {e}")

    # Enrich with live price
    doc = await _enrich_with_live_price(doc)
    try:
        if not doc.get("dcf_summary"):
            doc["dcf_summary"] = _load_dcf_summary(doc.get("ticker"))
    except Exception:
        pass
    return doc


@api_router.get("/research/{session_id}/swot")
async def get_swot(session_id: str):
    _empty = {"strengths": [], "weaknesses": [], "opportunities": [], "threats": []}
    try:
        if not RP_SWOT or rp_generate_swot is None:
            return {**_empty, "error": "SWOT engine unavailable"}
        rp_ses = _find_rp_session(session_id)
        if rp_ses is None:
            db = _get_db()
            if db is not None:
                try:
                    doc = await db.sessions.find_one({"_id": session_id})
                    if doc and doc.get("ticker"):
                        rp_ses = _find_rp_session_by_ticker(doc["ticker"])
                except Exception:
                    pass
        if rp_ses is None:
            return {**_empty, "error": "Session not found"}
        sector = "other"
        try:
            meta = rp_ses.get_meta() if hasattr(rp_ses, "get_meta") else {}
            sector = meta.get("sector") or _detect_sector_simple(getattr(rp_ses, "ticker", "")) or "other"
        except Exception:
            pass
        return rp_generate_swot(rp_ses, None, sector=sector)
    except Exception as exc:
        logger.warning(f"[swot] {session_id}: {exc}")
        return {**_empty, "error": str(exc)}


@api_router.get("/research/{session_id}/porter")
async def get_porter(session_id: str):
    _empty = {"forces": []}
    try:
        if not RP_PORTER or rp_generate_porter is None:
            return {**_empty, "error": "Porter engine unavailable"}
        rp_ses = _find_rp_session(session_id)
        if rp_ses is None:
            db = _get_db()
            if db is not None:
                try:
                    doc = await db.sessions.find_one({"_id": session_id})
                    if doc and doc.get("ticker"):
                        rp_ses = _find_rp_session_by_ticker(doc["ticker"])
                except Exception:
                    pass
        if rp_ses is None:
            return {**_empty, "error": "Session not found"}
        sector = "other"
        try:
            meta = rp_ses.get_meta() if hasattr(rp_ses, "get_meta") else {}
            sector = meta.get("sector") or _detect_sector_simple(getattr(rp_ses, "ticker", "")) or "other"
        except Exception:
            pass
        raw = rp_generate_porter(rp_ses, sector=sector)
        if isinstance(raw, list):
            return {"forces": raw}
        if isinstance(raw, dict):
            _name_map = {
                "competitive_rivalry":    "Competitive Rivalry",
                "supplier_power":         "Supplier Power",
                "buyer_power":            "Buyer Power",
                "threat_of_substitutes":  "Threat of Substitutes",
                "threat_of_new_entrants": "Threat of New Entrants",
            }
            forces = [
                {"name": _name_map.get(k, k.replace("_", " ").title()), **v}
                for k, v in raw.items()
            ]
            return {"forces": forces}
        return _empty
    except Exception as exc:
        logger.warning(f"[porter] {session_id}: {exc}")
        return {**_empty, "error": str(exc)}


# ═══════════════════════════════════════════════════════════════════════════
# DCF NOTEBOOK EXECUTION — async, papermill-based
# ═══════════════════════════════════════════════════════════════════════════

# In-memory per-session DCF run state
# {session_id: {"status": str, "started_at": datetime, "completed_at": datetime|None,
#               "elapsed": float, "error": str|None, "output_path": str|None, "ticker": str}}
_DCF_RUNS: dict[str, dict] = {}
_DCF_LOCK = asyncio.Lock()

# Paths
DCF_NOTEBOOK_SRC = Path("/app/notebooks/DCF_Multi_Source_Pipeline_REFACTORED.ipynb")


def _get_dcf_excel_p30(ticker: str) -> float | None:
    """Read Implied Share Price from cell P30 (DCF sheet) of the DCF Excel workbook.

    P30 is a formula =P29/R25.  openpyxl data_only mode may return None when
    cached formula values are absent, so we evaluate the full formula chain
    manually using the cached scalar constants that ARE present.
    """
    t = ticker.upper().strip()
    ticker_ns = t if "." in t else f"{t}.NS"
    excel_path = Path("/app/notebooks") / f"DCF_Output_{ticker_ns}_INR.xlsm"
    if not excel_path.exists():
        logger.warning(f"[dcf_excel_p30] {ticker}: Excel not found at {excel_path}")
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        sheet_name = next((sn for sn in wb.sheetnames if sn.strip() == "DCF"), None)
        if sheet_name is None:
            logger.warning(f"[dcf_excel_p30] {ticker}: no DCF sheet in {wb.sheetnames}")
            return None
        ws = wb[sheet_name]

        # Fast path: cached formula result
        p30 = ws["P30"].value
        if p30 is not None and isinstance(p30, (int, float)):
            return float(p30)

        # Slow path: evaluate formula chain from cached scalar inputs
        def _f(cell_val) -> float:
            return float(cell_val) if cell_val is not None else 0.0

        G6  = _f(ws["G6"].value)
        R10 = _f(ws["R10"].value)
        R11 = _f(ws["R11"].value)
        R12 = _f(ws["R12"].value)
        R14 = _f(ws["R14"].value)
        R15 = _f(ws["R15"].value)
        R16 = _f(ws["R16"].value)
        R18 = _f(ws["R18"].value)
        R25 = _f(ws["R25"].value)
        P25 = _f(ws["P25"].value)
        P26 = _f(ws["P26"].value)

        if R25 == 0:
            return None

        # Stub period: R6 = (R7 - R8) / 365
        R7 = ws["R7"].value
        R8 = ws["R8"].value
        if not isinstance(R7, datetime) or not isinstance(R8, datetime):
            logger.warning(f"[dcf_excel_p30] {ticker}: R7/R8 not datetime ({R7!r}, {R8!r})")
            return None
        stub = (R7 - R8).days / 365.0

        # Cost of Capital sheet
        ws_coc = wb.get("Cost of Capital")
        if ws_coc is None:
            return None
        C3  = _f(ws_coc["C3"].value)
        C4  = _f(ws_coc["C4"].value)
        C5  = _f(ws_coc["C5"].value)
        C6  = _f(ws_coc["C6"].value)
        C9  = _f(ws_coc["C9"].value)
        C10 = _f(ws_coc["C10"].value)
        C11 = _f(ws_coc["C11"].value)
        E27 = _f(ws_coc["E27"].value)
        P24 = C9  # P24 = CoC!C9 (debt)

        # Cash: P27 = Balance Sheet!F4 = Cash Flow!F29
        ws_cf = wb.get("Cash Flow")
        if ws_cf is None:
            return None
        P27 = _f(ws_cf["F29"].value)

        # WACC
        C12 = C9 + C10 + C11
        if C12 == 0:
            return None
        C15 = C9 / C12
        C16 = C10 / C12
        D20 = (E27 * C15) * (1 - C3)
        D21 = (C5 + C4 * C6) * C16
        WACC = D20 + D21
        if WACC <= R18:
            logger.warning(f"[dcf_excel_p30] {ticker}: WACC ({WACC}) <= terminal growth ({R18})")
            return None

        # 5-year projected FCFFs
        prev_rev = G6
        fcffs = []
        disc_periods = []
        for i in range(5):
            rev   = prev_rev * (1 + R10)
            ebit  = rev * R11
            nopat = ebit * (1 - R12)
            da    = rev * R14
            capex = -rev * R15
            wc    = (prev_rev - rev) * R16
            fcffs.append(nopat + da + capex + wc)
            disc_periods.append(stub + i)
            prev_rev = rev

        P13 = sum(f / (1 + WACC) ** n for f, n in zip(fcffs, disc_periods))

        # Terminal value (Gordon Growth)
        P6 = (fcffs[-1] * (1 + R18)) / (WACC - R18)
        P9 = P6 / (1 + WACC) ** disc_periods[-1]

        P23 = P9 + P13                       # Implied TEV
        P29 = P23 - P24 - P25 - P26 + P27   # Implied Equity Value
        result = P29 / R25
        logger.info(f"[dcf_excel_p30] {ticker}: P30={result:.4f} (TEV={P23:.0f}, WACC={WACC:.4f})")
        return result
    except Exception as _e:
        logger.warning(f"[dcf_excel_p30] {ticker}: {_e}")
        return None


def _load_dcf_summary(ticker: str) -> dict | None:
    """
    Load the papermill-written DCF summary JSON for a ticker and normalize
    into { forecast: [{year,revenue,ebit,fcff,revenue_growth}], meta: {...},
    sensitivity_table: null, ... }. Never raises.
    """
    if not ticker:
        return None
    t = ticker.upper().strip()
    ticker_ns = t if "." in t else f"{t}.NS"
    candidates = [
        Path("/app/notebooks") / f"DCF_Output_{ticker_ns}_INR.summary.json",
        DCF_NOTEBOOK_SRC.parent / f"DCF_Output_{ticker_ns}_INR.summary.json",
        Path(__file__).parent.parent / "notebooks" / f"DCF_Output_{ticker_ns}_INR.summary.json",
    ]
    for p in candidates:
        try:
            if p.exists():
                raw = json.loads(p.read_text(encoding="utf-8"))
                # Normalize: merge forecast_revenue + forecast_fcf into a
                # single `forecast` list with year/revenue/fcff/revenue_growth.
                rev_rows = raw.get("forecast_revenue") or []
                fcf_rows = raw.get("forecast_fcf") or []
                rev_by_fy = {int(r["fy"]): r for r in rev_rows if r.get("fy") is not None}
                fcf_by_fy = {int(r["fy"]): r for r in fcf_rows if r.get("fy") is not None}
                fys = sorted(set(rev_by_fy) | set(fcf_by_fy))
                forecast = []
                prev_rev = None
                for fy in fys:
                    r = rev_by_fy.get(fy, {})
                    f = fcf_by_fy.get(fy, {})
                    rev = r.get("revenue")
                    growth = None
                    if prev_rev not in (None, 0) and rev is not None:
                        try:
                            growth = (float(rev) / float(prev_rev) - 1.0)
                        except Exception:
                            growth = None
                    forecast.append({
                        "year": fy,
                        "revenue": rev,
                        "ebit": r.get("ebit"),
                        "fcff": f.get("fcff"),
                        "revenue_growth": growth,
                    })
                    if rev is not None:
                        prev_rev = rev
                raw["forecast"] = forecast
                raw["meta"] = {
                    "base_year": (raw.get("assumptions") or {}).get("base_year"),
                    "currency": raw.get("currency", "INR"),
                    "units": "millions",
                }
                return raw
        except Exception as e:
            logger.debug(f"_load_dcf_summary failed for {p}: {e}")
    return None
DCF_RUN_ROOT = Path("/tmp/dcf_runs")
DCF_RUN_ROOT.mkdir(parents=True, exist_ok=True)

# Max wait for a notebook run (seconds)
DCF_TIMEOUT_SECONDS = 180


def _dcf_run_dir(session_id: str) -> Path:
    d = DCF_RUN_ROOT / session_id
    d.mkdir(parents=True, exist_ok=True)
    return d


def _execute_notebook_blocking(
    session_id: str,
    ticker: str,
    timeout: int = DCF_TIMEOUT_SECONDS,
) -> dict:
    """
    Run the DCF notebook for a ticker using papermill.
    Writes to /tmp/dcf_runs/{session_id}/.
    This function is BLOCKING — call it from a background thread.

    Returns: {status, elapsed, output_path, error}
    """
    start = datetime.now(timezone.utc)
    run_dir = _dcf_run_dir(session_id)

    # Clean ticker (strip trailing spaces the notebook has in its default)
    ticker_clean = ticker.strip().upper()
    ticker_ns = ticker_clean if "." in ticker_clean else f"{ticker_clean}.NS"

    # Output notebook (executed copy) — papermill writes the executed version here
    executed_nb = run_dir / f"{ticker_clean}_executed.ipynb"

    # Result dict
    result = {
        "status": "running",
        "ticker": ticker_clean,
        "started_at": start.isoformat(),
        "output_path": None,
        "error": None,
        "elapsed": 0.0,
    }

    try:
        import papermill as pm

        if not DCF_NOTEBOOK_SRC.exists():
            raise FileNotFoundError(f"Notebook not found: {DCF_NOTEBOOK_SRC}")

        # Parameters to inject via papermill. These get written into a new cell
        # at the top of the notebook, which runs BEFORE Cell 2. Cell 2 has:
        #   TICKER_INPUT = "RELIANCE.NS "
        # which would overwrite our injected value. So we override differently
        # after-the-fact via the execute_notebook `parameters` dict.
        #
        # Strategy: papermill injects parameters BEFORE Cell 2 runs. Cell 2 then
        # reassigns TICKER_INPUT. To shadow it, we use a wrapper notebook
        # approach: we rewrite Cell 2 in a copy of the notebook so
        # TICKER_INPUT uses globals().get() like the other params already do.

        # Read notebook, patch Cell 2, save to run dir
        import nbformat
        nb_content = nbformat.read(str(DCF_NOTEBOOK_SRC), as_version=4)

        patched_cell_source = None
        for cell in nb_content.cells:
            if cell.cell_type != "code":
                continue
            if "TICKER_INPUT" in cell.source and "MASTER CONFIG" in cell.source:
                # Replace the hard-coded TICKER_INPUT line with a globals()-aware version
                lines = cell.source.split("\n")
                new_lines = []
                for line in lines:
                    stripped = line.lstrip()
                    if stripped.startswith("TICKER_INPUT =") and "globals" not in line:
                        # Preserve indentation
                        indent = line[:len(line) - len(stripped)]
                        new_lines.append(
                            f'{indent}TICKER_INPUT = globals().get("TICKER_INPUT") or "{ticker_ns}"'
                        )
                    else:
                        new_lines.append(line)
                cell.source = "\n".join(new_lines)
                patched_cell_source = cell.source
                break

        if patched_cell_source is None:
            raise RuntimeError("Could not find TICKER_INPUT cell to patch")

        # Save patched notebook
        patched_nb = run_dir / "patched_input.ipynb"
        nbformat.write(nb_content, str(patched_nb))

        logger.info(f"[dcf] Running notebook for {ticker_clean} (session {session_id})")
        logger.info(f"[dcf]   patched notebook: {patched_nb}")
        logger.info(f"[dcf]   executed output:  {executed_nb}")

        # Execute via papermill
        # Working directory = notebooks folder so relative paths work
        # Note: papermill 2.7+ doesn't accept stderr_file as str - use logging instead
        pm.execute_notebook(
            input_path=str(patched_nb),
            output_path=str(executed_nb),
            parameters={"TICKER_INPUT": ticker_ns},
            kernel_name="python3",
            cwd=str(DCF_NOTEBOOK_SRC.parent),
            progress_bar=False,
            log_output=True,  # emit cell outputs to logger instead of stderr_file
            execution_timeout=timeout,
        )

        # The notebook writes output to: notebooks/DCF_Output_{TICKER}_{CCY}.xls[mx]
        # Where TICKER may be bare (RELIANCE) or exchange-suffixed (RELIANCE.NS).
        # Extension may be .xlsx or .xlsm (template has macros, so .xlsm is common).
        bare_ticker = ticker_clean.replace(".NS", "").replace(".BO", "")

        # Try multiple patterns, most-specific first
        candidates = []
        for pattern in [
            f"DCF_Output_{ticker_ns}_*.xlsm",   # RELIANCE.NS_INR.xlsm ← actual
            f"DCF_Output_{ticker_ns}_*.xlsx",
            f"DCF_Output_{bare_ticker}_*.xlsm",
            f"DCF_Output_{bare_ticker}_*.xlsx",
            f"DCF_Output_*{bare_ticker}*.xlsm",
            f"DCF_Output_*{bare_ticker}*.xlsx",
        ]:
            hits = list(DCF_NOTEBOOK_SRC.parent.glob(pattern))
            if hits:
                candidates = hits
                break

        if not candidates:
            # List everything that starts with DCF_Output_ so the error is useful
            all_outputs = list(DCF_NOTEBOOK_SRC.parent.glob("DCF_Output_*"))
            raise FileNotFoundError(
                f"Notebook ran but no DCF_Output_{bare_ticker}_*.xls[mx] found in "
                f"{DCF_NOTEBOOK_SRC.parent}. Files present: {[p.name for p in all_outputs]}"
            )

        # Pick the newest
        output_file = max(candidates, key=lambda p: p.stat().st_mtime)

        # Copy to session-specific folder so we keep it even if notebook re-runs
        session_output = run_dir / output_file.name
        import shutil
        shutil.copy2(output_file, session_output)

        elapsed = (datetime.now(timezone.utc) - start).total_seconds()
        result.update({
            "status": "complete",
            "output_path": str(session_output),
            "output_filename": output_file.name,
            "elapsed": round(elapsed, 1),
            "completed_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info(f"[dcf] ✓ {ticker_clean} complete in {elapsed:.1f}s → {output_file.name}")

    except Exception as e:
        elapsed = (datetime.now(timezone.utc) - start).total_seconds()
        err_detail = str(e)[:500]

        # Try to extract the actual cell error from the partially-executed notebook
        try:
            if executed_nb.exists():
                import nbformat
                nb_err = nbformat.read(str(executed_nb), as_version=4)
                for idx, cell in enumerate(nb_err.cells):
                    if cell.cell_type != "code":
                        continue
                    for out in cell.get("outputs", []):
                        if out.get("output_type") == "error":
                            tb = "\n".join(out.get("traceback", []))
                            # Strip ANSI escape codes
                            import re as _re
                            tb = _re.sub(r"\x1b\[[0-9;]*[mK]", "", tb)
                            err_detail = (
                                f"{str(e)[:200]}\n"
                                f"---failing cell {idx}---\n"
                                f"{tb[-1500:]}"
                            )
                            break
                    if "failing cell" in err_detail:
                        break
        except Exception as _inner:
            logger.debug(f"[dcf] could not extract cell error: {_inner}")

        result.update({
            "status": "failed",
            "error": err_detail,
            "elapsed": round(elapsed, 1),
            "completed_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.error(f"[dcf] ✗ {ticker_clean} failed after {elapsed:.1f}s: {e}")

    return result


async def _run_dcf_background(session_id: str, ticker: str):
    """Wrapper: runs the blocking notebook execution in a thread."""
    loop = asyncio.get_event_loop()

    # Update state to running
    async with _DCF_LOCK:
        _DCF_RUNS[session_id] = {
            "status": "running",
            "ticker": ticker,
            "started_at": datetime.now(timezone.utc).isoformat(),
            "elapsed": 0.0,
            "error": None,
            "output_path": None,
        }

    # Execute in thread (papermill is blocking)
    try:
        result = await loop.run_in_executor(
            None,
            _execute_notebook_blocking,
            session_id,
            ticker,
        )
    except Exception as e:
        result = {
            "status": "failed",
            "ticker": ticker,
            "error": f"background executor failed: {e}",
            "elapsed": 0.0,
            "output_path": None,
        }

    # Store final result
    async with _DCF_LOCK:
        _DCF_RUNS[session_id] = result


@api_router.post("/research/{session_id}/dcf/run")
@safe_endpoint(lambda: {"status": "error", "error": "failed to kick off DCF run"})
async def dcf_run(session_id: str, request: Request):
    """
    Kick off the DCF notebook for this session's ticker.
    Returns immediately with status=running. Poll /dcf/status for progress.
    """
    # Get ticker from request body OR from session
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    ticker = (body.get("ticker") or "").strip().upper()

    if not ticker:
        # Look up ticker from Mongo session
        db = _get_db()
        if db is not None:
            try:
                doc = await db.sessions.find_one({"_id": session_id})
                if doc:
                    ticker = (doc.get("ticker") or "").strip().upper()
            except Exception:
                pass

    if not ticker:
        # Try cached session
        cached, _ = await cache_get(f"session:{session_id}")
        if cached:
            ticker = (cached.get("ticker") or "").strip().upper()

    if not ticker:
        raise HTTPException(status_code=400, detail="could not resolve ticker for session")

    # Check if already running
    async with _DCF_LOCK:
        existing = _DCF_RUNS.get(session_id)
        if existing and existing.get("status") == "running":
            return {
                "status": "already_running",
                "session_id": session_id,
                "ticker": existing.get("ticker"),
                "started_at": existing.get("started_at"),
                "message": "DCF already running for this session",
            }

    # Kick off background task
    asyncio.create_task(_run_dcf_background(session_id, ticker))

    return {
        "status": "running",
        "session_id": session_id,
        "ticker": ticker,
        "started_at": datetime.now(timezone.utc).isoformat(),
        "timeout_seconds": DCF_TIMEOUT_SECONDS,
        "poll_url": f"/api/research/{session_id}/dcf/status",
    }


@api_router.get("/research/{session_id}/dcf/status")
@safe_endpoint(lambda: {"status": "idle"})
async def dcf_status(session_id: str):
    """Poll the status of a running DCF notebook."""
    async with _DCF_LOCK:
        state = _DCF_RUNS.get(session_id)

    if not state:
        return {"status": "idle", "session_id": session_id}

    # Compute elapsed if still running
    if state.get("status") == "running":
        try:
            started = datetime.fromisoformat(state["started_at"])
            elapsed = (datetime.now(timezone.utc) - started).total_seconds()
            state = {**state, "elapsed": round(elapsed, 1)}
        except Exception:
            pass

    # Choose download URL matching actual output extension
    download_url = None
    if state.get("status") == "complete":
        output_filename = state.get("output_filename", "") or ""
        if output_filename.lower().endswith(".xlsm"):
            download_url = f"/api/research/{session_id}/dcf/output.xlsm"
        else:
            download_url = f"/api/research/{session_id}/dcf/output.xlsx"

    return {
        "session_id": session_id,
        **state,
        # Don't leak local filesystem path to frontend
        "output_path": None,
        "download_url": download_url,
    }


@api_router.get("/research/{session_id}/dcf/output.xlsx")
@api_router.get("/research/{session_id}/dcf/output.xlsm")
async def dcf_output_download(session_id: str):
    """Download the DCF notebook's output xlsx/xlsm for a completed run."""
    from fastapi.responses import Response

    async with _DCF_LOCK:
        state = _DCF_RUNS.get(session_id)

    if not state or state.get("status") != "complete":
        # Try to find output on disk (maybe backend restarted after run)
        run_dir = _dcf_run_dir(session_id)
        candidates = sorted(
            list(run_dir.glob("DCF_Output_*.xlsm")) + list(run_dir.glob("DCF_Output_*.xlsx")),
            key=lambda p: p.stat().st_mtime,
        )
        if not candidates:
            raise HTTPException(
                status_code=404,
                detail="No completed DCF run for this session. POST to /dcf/run first.",
            )
        output_path = candidates[-1]
    else:
        output_path = Path(state["output_path"])

    if not output_path.exists():
        raise HTTPException(status_code=404, detail=f"Output file missing: {output_path.name}")

    # Use correct MIME: .xlsm has macros, different content type
    if output_path.suffix.lower() == ".xlsm":
        media_type = "application/vnd.ms-excel.sheet.macroEnabled.12"
    else:
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    content = output_path.read_bytes()
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{output_path.name}"'},
    )


@api_router.delete("/research/{session_id}/dcf/cancel")
@safe_endpoint(lambda: {"status": "error"})
async def dcf_cancel(session_id: str):
    """Clear DCF run state for this session (doesn't kill running notebook)."""
    async with _DCF_LOCK:
        _DCF_RUNS.pop(session_id, None)
    return {"status": "cleared", "session_id": session_id}


# Research sub-endpoints — all return safe empty payloads
@api_router.post("/research/{session_id}/run-scenarios")
@safe_endpoint(lambda: {"scenarios": [], "_fallback": True})
async def run_scenarios(session_id: str):
    return {"session_id": session_id, "scenarios": []}

@api_router.post("/research/{session_id}/catalyst")
@safe_endpoint(lambda: {"catalysts": [], "_fallback": True})
async def add_catalyst(session_id: str, request: Request):
    return {"session_id": session_id, "catalysts": []}

@api_router.post("/research/{session_id}/thesis")
@safe_endpoint(lambda: {"thesis": "", "_fallback": True})
async def update_thesis(session_id: str, request: Request):
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    return {"session_id": session_id, "thesis": body.get("thesis", "")}

@api_router.post("/research/{session_id}/dcf")
@safe_endpoint(lambda: {"dcf": None, "_fallback": True})
async def run_dcf(session_id: str, request: Request):
    return {"session_id": session_id, "dcf": None, "message": "DCF endpoint stub"}

@api_router.get("/research/{session_id}/dcf")
@safe_endpoint(lambda: {"dcf": None, "_fallback": True})
async def get_dcf(session_id: str):
    return {"session_id": session_id, "dcf": None}

# ---------- Report Downloads (HTML + XLSX) ----------

def _find_rp_session(session_id: str):
    """
    Find a research_platform session by various means:
      1. Direct session_id match
      2. Search filesystem sessions for matching rp_session_id in meta
      3. None if not found
    Returns ResearchSession or None. Never raises.
    """
    if not RP_SESSION_MGR or rp_load_session is None:
        return None

    # Strategy 1: direct load (session_id IS the RP session_id)
    try:
        return rp_load_session(session_id)
    except Exception:
        pass

    # Strategy 2: scan sessions folder for a matching ticker
    # (when Mongo session_id != RP session_id, e.g. fallback IDs)
    try:
        if rp_list_sessions is not None:
            all_sessions = rp_list_sessions()
            # Look for exact session_id match
            for s in all_sessions:
                if s.get("session_id") == session_id:
                    return rp_load_session(session_id)
    except Exception:
        pass

    return None


def _find_rp_session_by_ticker(ticker: str):
    """Find the most recent RP session for a given ticker. Returns None on failure."""
    if not RP_SESSION_MGR or rp_list_sessions is None or rp_load_session is None:
        return None
    try:
        sessions = rp_list_sessions(ticker=ticker)
        if sessions:
            return rp_load_session(sessions[0]["session_id"])
    except Exception:
        pass
    return None


async def _load_session_for_report(session_id: str) -> dict:
    """Fetch session data from Mongo or cache for report generation."""
    db = _get_db()
    if db is not None:
        try:
            doc = await db.sessions.find_one({"_id": session_id})
            if doc:
                doc["_id"] = str(doc.get("_id", ""))
                if isinstance(doc.get("created_at"), datetime):
                    doc["created_at"] = doc["created_at"].isoformat()
                # Enrich with cached analyze response for price/DCF
                cached, _ = await cache_get(f"session:{session_id}")
                if cached:
                    # Merge, prefer cached for numeric fields
                    for k in ("price", "ltp", "prev_close", "change", "change_percent", "dcf",
                              "scoring", "factor_scores"):
                        if cached.get(k) is not None:
                            doc[k] = cached[k]
                return doc
        except Exception as e:
            logger.debug(f"report load mongo: {e}")
    cached, _ = await cache_get(f"session:{session_id}")
    return cached or {"session_id": session_id, "ticker": "UNKNOWN", "status": "not_found"}


def _indian_format(n, digits=0):
    """Format number using Indian lakh/crore grouping (1,23,45,678)."""
    try:
        if n is None:
            return "—"
        num = float(n)
        if num != num:  # NaN
            return "—"
        neg = num < 0
        num = abs(num)
        if digits > 0:
            s = f"{num:.{digits}f}"
            int_part, dec_part = s.split(".")
        else:
            int_part = str(int(round(num)))
            dec_part = None
        if len(int_part) <= 3:
            grouped = int_part
        else:
            last3 = int_part[-3:]
            rest = int_part[:-3]
            pairs = []
            while len(rest) > 2:
                pairs.append(rest[-2:])
                rest = rest[:-2]
            if rest:
                pairs.append(rest)
            grouped = ",".join(reversed(pairs)) + "," + last3
        if dec_part:
            grouped = f"{grouped}.{dec_part}"
        return ("-" if neg else "") + grouped
    except Exception:
        return "—"


_yf_cache = {}
_YF_CACHE_TTL = 600

PEER_MAP = {
    "IT":              ["INFY", "WIPRO", "HCLTECH"],
    "TECHNOLOGY":      ["INFY", "WIPRO", "HCLTECH"],
    "SOFTWARE":        ["INFY", "WIPRO", "HCLTECH"],
    "BANKING":         ["HDFCBANK", "ICICIBANK", "AXISBANK"],
    "BANKS":           ["HDFCBANK", "ICICIBANK", "AXISBANK"],
    "FINANCIAL":       ["HDFCBANK", "ICICIBANK", "AXISBANK"],
    "PHARMA":          ["SUNPHARMA", "CIPLA", "DRREDDY"],
    "PHARMACEUTICALS": ["SUNPHARMA", "CIPLA", "DRREDDY"],
    "HEALTHCARE":      ["SUNPHARMA", "CIPLA", "DRREDDY"],
    "ENERGY":          ["RELIANCE", "ONGC", "IOC"],
    "OIL":             ["RELIANCE", "ONGC", "IOC"],
    "FMCG":            ["HINDUNILVR", "NESTLEIND", "ITC"],
    "CONSUMER":        ["HINDUNILVR", "NESTLEIND", "ITC"],
    "DEFAULT":         ["RELIANCE", "TCS", "INFY"],
}


def _yf_info(ticker: str) -> dict:
    """Cached yfinance lookup; 10-min TTL. Returns {} on failure."""
    import time as _time
    now = _time.time()
    cached = _yf_cache.get(ticker)
    if cached and now - cached[0] < _YF_CACHE_TTL:
        return cached[1]
    info = {}
    try:
        import yfinance as yf
        info = yf.Ticker(f"{ticker}.NS").info or {}
    except Exception as e:
        logger.warning(f"yfinance fetch failed for {ticker}: {e}")
    _yf_cache[ticker] = (now, info)
    return info


def _read_rp_session_file(rp_ses, filename: str) -> dict:
    """Read a JSON file from an RP session's directory. Returns {} on any failure."""
    if rp_ses is None:
        return {}
    try:
        sdir = (getattr(rp_ses, "session_dir", None)
                or getattr(rp_ses, "path", None)
                or getattr(rp_ses, "dir", None))
        if sdir is None:
            return {}
        fp = Path(sdir) / filename
        if not fp.exists():
            return {}
        import json as _json
        return _json.loads(fp.read_text(encoding="utf-8")) or {}
    except Exception as e:
        logger.debug(f"read_rp_session_file({filename}): {e}")
        return {}


async def build_report_context(session_id: str, ticker: str, session: dict) -> dict:
    """
    Gather report context from tiered sources. Never raises.
    Returns dict; missing fields are None.
    """
    ctx = {
        "thesis_line": None,
        "rationale_our_view": None,
        "rationale_catalyst": None,
        "rationale_valuation": None,
        "bull_case_tp": None,
        "bear_case_tp": None,
        "val_caption": None,
        "company_name": None,
        "sector": None,
        "cmp": None,
        "cmp_date": None,
        "market_cap_cr": None,
        "week52_high": None,
        "week52_low": None,
        "avg_volume_cr": None,
        "free_float_pct": None,
        "beta": None,
        "peers": [],
        "exec_summary": None,
    }
    tiers = {"tier1": False, "tier2": False, "tier3": False,
             "tier4": False, "tier5": False, "tier6": False}

    # ── TIER 1: research_platform session files ───────────────────────
    rp_ses = None
    try:
        rp_ses = _find_rp_session(session_id) or _find_rp_session_by_ticker(ticker)
    except Exception:
        pass
    if rp_ses is not None:
        try:
            meta = _read_rp_session_file(rp_ses, "session_meta.json")
            if isinstance(meta, dict):
                if meta.get("thesis"):
                    ctx["thesis_line"] = str(meta["thesis"]).strip()
                if meta.get("variant_view"):
                    ctx["rationale_our_view"] = str(meta["variant_view"]).strip()
                cats = meta.get("catalysts") or []
                if isinstance(cats, list) and cats:
                    first = cats[0]
                    desc = first.get("description") if isinstance(first, dict) else str(first)
                    if desc:
                        ctx["rationale_catalyst"] = str(desc).strip()
            scen = _read_rp_session_file(rp_ses, "scenarios.json")
            scen_data = scen.get("scenarios") if isinstance(scen, dict) else None
            scen_data = scen_data or (scen if isinstance(scen, dict) else {})
            base = scen_data.get("base") or {}
            bull = scen_data.get("bull") or {}
            bear = scen_data.get("bear") or {}
            if isinstance(base, dict) and base.get("key_assumption"):
                ctx["rationale_valuation"] = str(base["key_assumption"]).strip()
            ctx["bull_case_tp"] = (bull.get("price_per_share") or bull.get("per_share")
                                   if isinstance(bull, dict) else None)
            ctx["bear_case_tp"] = (bear.get("price_per_share") or bear.get("per_share")
                                   if isinstance(bear, dict) else None)
            assump = _read_rp_session_file(rp_ses, "assumptions.json")
            if isinstance(assump, dict):
                wacc = assump.get("wacc")
                tg = assump.get("terminal_growth_rate") or assump.get("terminal_growth")
                if wacc is not None and tg is not None:
                    try:
                        ctx["val_caption"] = (
                            f"WACC {float(wacc)*100:.1f}%, "
                            f"terminal growth {float(tg)*100:.1f}%"
                        )
                    except Exception:
                        pass
            tiers["tier1"] = any([ctx["thesis_line"], ctx["rationale_our_view"],
                                  ctx["rationale_catalyst"], ctx["rationale_valuation"],
                                  ctx["bull_case_tp"], ctx["bear_case_tp"]])
        except Exception as e:
            logger.warning(f"tier1 report data failed for {ticker}: {e}")

    # ── TIER 2: MongoDB company + price_history ───────────────────────
    try:
        db = _get_db()
        if db is not None:
            try:
                company = await db["company"].find_one({"ticker": ticker})
                if company:
                    ctx["company_name"] = company.get("name") or ctx["company_name"]
                    ctx["sector"] = (company.get("sector_mapped")
                                     or company.get("sector") or ctx["sector"])
                    tiers["tier2"] = True
            except Exception as e:
                logger.debug(f"company lookup failed: {e}")
            try:
                ph = await db["price_history"].find_one(
                    {"ticker": ticker}, sort=[("date", -1)]
                )
                if ph:
                    ctx["cmp"] = ph.get("close") or ctx["cmp"]
                    d = ph.get("date")
                    if d is not None:
                        if isinstance(d, datetime):
                            ctx["cmp_date"] = d.strftime("%d %b %Y")
                        else:
                            ctx["cmp_date"] = str(d)
                    tiers["tier2"] = True
            except Exception as e:
                logger.debug(f"price_history lookup failed: {e}")
    except Exception as e:
        logger.warning(f"tier2 report data failed for {ticker}: {e}")

    # ── TIER 3: yfinance ──────────────────────────────────────────────
    try:
        info = _yf_info(ticker)
        if info:
            mc = info.get("marketCap")
            if mc:
                try: ctx["market_cap_cr"] = float(mc) / 1e7
                except Exception: pass
            ctx["week52_high"] = info.get("fiftyTwoWeekHigh") or ctx["week52_high"]
            ctx["week52_low"] = info.get("fiftyTwoWeekLow") or ctx["week52_low"]
            av = info.get("averageDailyVolume10Day")
            if av:
                try:
                    price_for_vol = info.get("currentPrice") or info.get("regularMarketPrice") or ctx["cmp"] or 0
                    if price_for_vol:
                        ctx["avg_volume_cr"] = float(av) * float(price_for_vol) / 1e7
                except Exception: pass
            fl = info.get("floatShares")
            so = info.get("sharesOutstanding")
            if fl and so:
                try: ctx["free_float_pct"] = float(fl) / float(so) * 100
                except Exception: pass
            if info.get("beta") is not None:
                ctx["beta"] = info["beta"]
            if ctx["cmp"] is None:
                ctx["cmp"] = info.get("currentPrice") or info.get("regularMarketPrice")
            tiers["tier3"] = any([mc, ctx["week52_high"], ctx["week52_low"], av, ctx["beta"]])
    except Exception as e:
        logger.warning(f"tier3 report data failed for {ticker}: {e}")

    # ── TIER 4: summary.md exec summary ───────────────────────────────
    try:
        if rp_ses is not None:
            sdir = (getattr(rp_ses, "session_dir", None)
                    or getattr(rp_ses, "path", None)
                    or getattr(rp_ses, "dir", None))
            if sdir:
                sfp = Path(sdir) / "summary.md"
                if sfp.exists():
                    txt = sfp.read_text(encoding="utf-8").strip()
                    if txt:
                        para = txt.split("\n\n")[0].strip()
                        ctx["exec_summary"] = para
                        tiers["tier4"] = True
    except Exception as e:
        logger.debug(f"tier4 summary.md read failed: {e}")
    if not ctx["exec_summary"]:
        ctx["exec_summary"] = "Detailed analysis available in the Excel model download."

    # ── Peers ─────────────────────────────────────────────────────────
    try:
        sec = (ctx["sector"] or session.get("sector") or "").upper()
        peer_key = None
        for k in PEER_MAP:
            if k != "DEFAULT" and k in sec:
                peer_key = k
                break
        peers = PEER_MAP.get(peer_key) if peer_key else PEER_MAP["DEFAULT"]
        ctx["peers"] = [p for p in peers if p.upper() != ticker.upper()][:3]
    except Exception:
        pass

    ctx["_tiers"] = tiers
    logger.info(
        f"Report data for {ticker}: "
        f"session={tiers['tier1']}, db={tiers['tier2']}, "
        f"yf={tiers['tier3']}, llm={tiers['tier4']}, "
        f"swot={tiers['tier5']}, financials={tiers['tier6']}"
    )
    return ctx


_report_llm_cache: dict = {}

_REPORT_EMPTY_CONTENT = {
    "thesis": "", "exec_summary": "",
    "rationale_consensus": "", "rationale_our_view": "",
    "rationale_catalyst": "", "rationale_valuation": "",
    "driver_1_name": "", "driver_1_desc": "",
    "driver_2_name": "", "driver_2_desc": "",
    "driver_3_name": "", "driver_3_desc": "",
    "headwind_1_name": "", "headwind_1_desc": "",
    "headwind_2_name": "", "headwind_2_desc": "",
    "net_view": "",
    "opp_1_name": "", "opp_1_desc": "",
    "opp_2_name": "", "opp_2_desc": "",
    "risk_1_name": "", "risk_1_desc": "",
    "risk_2_name": "", "risk_2_desc": "",
    "kill_switch": "",
}


def generate_report_commentary(session_id: str, ticker: str, sector: str,
                               ctx: dict, session: dict) -> dict:
    """
    Use Claude to generate institutional report commentary from real DCF data.
    Cached per session_id. Never raises — returns empty-string dict on failure.
    """
    if session_id in _report_llm_cache:
        return _report_llm_cache[session_id]

    if not ANTHROPIC_KEY:
        logger.warning("[report-llm] ANTHROPIC_API_KEY missing; skipping LLM commentary")
        return dict(_REPORT_EMPTY_CONTENT)

    # Assemble real numbers from whatever tiers landed
    base_price = None
    scen = session.get("scenarios") or {}
    if isinstance(scen, dict):
        base = scen.get("base") or {}
        if isinstance(base, dict):
            base_price = base.get("price_per_share") or base.get("per_share")
    if base_price is None:
        base_price = (session.get("dcf") or {}).get("fair_value")

    rating = None
    try:
        rating = ((scen.get("base") or {}).get("rating")
                  or (session.get("scoring") or {}).get("recommendation"))
    except Exception:
        pass

    upside = None
    try:
        upside = (scen.get("base") or {}).get("upside_pct")
    except Exception:
        pass

    wacc = (session.get("assumptions") or {}).get("wacc")
    tg = ((session.get("assumptions") or {}).get("terminal_growth_rate")
          or (session.get("assumptions") or {}).get("terminal_growth"))
    current_price = ctx.get("cmp") or session.get("current_price")

    def _fmt(v, pct=False, d=1):
        if v is None:
            return "n/a"
        try:
            if pct:
                return f"{float(v)*100:.{d}f}%"
            return f"{float(v):.0f}"
        except Exception:
            return "n/a"

    prompt = f"""You are writing an institutional equity research report for {ticker} ({sector or 'Indian equities'} sector). Here is the DCF output from our quantitative model:

- Base case target price: ₹{_fmt(base_price)}
- Bull case target: ₹{_fmt(ctx.get('bull_case_tp'))}
- Bear case target: ₹{_fmt(ctx.get('bear_case_tp'))}
- Current market price: ₹{_fmt(current_price)}
- Upside: {_fmt(upside, d=1) if upside is not None else 'n/a'}%
- Rating: {rating or 'n/a'}
- WACC used: {_fmt(wacc, pct=True)}
- Terminal growth: {_fmt(tg, pct=True)}

Generate the following report sections as a JSON object. Be specific, data-grounded, and institutional in tone. 50-80 words per section max.

{{
  "thesis": "10-15 word thesis line. What's the differentiated view?",
  "exec_summary": "3-sentence executive summary. Open with rating+target, explain thesis, name the catalyst.",
  "rationale_consensus": "1 sentence describing what consensus believes.",
  "rationale_our_view": "1 sentence describing our differentiated view.",
  "rationale_catalyst": "1 sentence describing the catalyst.",
  "rationale_valuation": "1 sentence on valuation anchor.",
  "driver_1_name": "3-4 word name of positive driver",
  "driver_1_desc": "1-sentence quantified driver description.",
  "driver_2_name": "3-4 word name",
  "driver_2_desc": "1-sentence driver description.",
  "driver_3_name": "3-4 word name",
  "driver_3_desc": "1-sentence driver description.",
  "headwind_1_name": "3-4 word name",
  "headwind_1_desc": "1-sentence headwind description.",
  "headwind_2_name": "3-4 word name",
  "headwind_2_desc": "1-sentence headwind description.",
  "net_view": "1-sentence net view for why positives outweigh.",
  "opp_1_name": "3-4 word upside optionality",
  "opp_1_desc": "1 sentence with probability and TP impact.",
  "opp_2_name": "3-4 word upside optionality",
  "opp_2_desc": "1 sentence with probability and TP impact.",
  "risk_1_name": "3-4 word downside risk",
  "risk_1_desc": "1 sentence with probability and TP impact.",
  "risk_2_name": "3-4 word downside risk",
  "risk_2_desc": "1 sentence with probability and TP impact.",
  "kill_switch": "1 sentence with specific observable data point."
}}

Return ONLY the JSON object, no markdown, no commentary."""

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_KEY, timeout=15.0)
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}],
        )
        text = resp.content[0].text.strip()
        # Strip possible ```json fences
        if text.startswith("```"):
            text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text, flags=re.DOTALL)
        data = json.loads(text)
        out = dict(_REPORT_EMPTY_CONTENT)
        for k in out:
            if k in data and isinstance(data[k], str):
                out[k] = data[k].strip()
        _report_llm_cache[session_id] = out
        logger.info(f"[report-llm] generated commentary for {ticker} ({session_id})")
        return out
    except Exception as e:
        logger.warning(f"[report-llm] generation failed for {ticker}: {e}")
        return dict(_REPORT_EMPTY_CONTENT)


@api_router.get("/research/{session_id}/report/download")
async def report_download_html(session_id: str):
    """
    Return institutional Beaver ER HTML report as a download.
    Populates the Beaver_ER_Template_Pro_Forest.html template with session data.
    Missing fields render as "—"; never returns 500.
    """
    from fastapi.responses import HTMLResponse, Response
    from html import escape as _html_escape
    import re

    # ─── Template-based report ────────────────────────────────────────────
    try:
        template_path = Path(__file__).parent / "templates" / "reports" / "Beaver_ER_Template_Pro_Forest.html"
        if template_path.exists():
            session = await _load_session_for_report(session_id)
            ticker = (session.get("ticker") or "UNKNOWN").upper()
            meta = session.get("meta") or {}
            long_name = meta.get("long_name") or session.get("long_name") or ticker
            sector = session.get("sector") or meta.get("sector") or "—"

            try:
                ctx = await build_report_context(session_id, ticker, session)
            except Exception as _e:
                logger.warning(f"[report] build_report_context failed: {_e}")
                ctx = {}
            if ctx.get("company_name"):
                long_name = ctx["company_name"]
            if ctx.get("sector") and ctx["sector"] != "—":
                sector = ctx["sector"]

            current_price = (
                ctx.get("cmp")
                or session.get("current_price")
                or session.get("price")
                or (session.get("dcf") or {}).get("current_price")
            )
            dcf = session.get("dcf") or {}
            scoring = session.get("scoring") or {}
            scenarios = session.get("scenarios") or {}

            # Target price / rating / upside — prefer base scenario, fall back to dcf
            base = scenarios.get("base") if isinstance(scenarios, dict) else None
            target = (base or {}).get("per_share") if base else dcf.get("fair_value")
            upside = (base or {}).get("upside_pct") if base else dcf.get("upside_pct")
            rating = ((base or {}).get("rating") or scoring.get("recommendation") or "HOLD")
            if not rating or str(rating).strip() in ("", "—"):
                rating = "HOLD"

            market_cap_cr = None
            mc_raw = meta.get("market_cap") or session.get("market_cap") or dcf.get("equity_value")
            if mc_raw is not None:
                try:
                    market_cap_cr = float(mc_raw) / 1e7  # rupees → crores
                except Exception:
                    market_cap_cr = None

            beta = (session.get("assumptions") or {}).get("beta") or meta.get("beta") or ctx.get("beta")
            if ctx.get("market_cap_cr"):
                market_cap_cr = ctx["market_cap_cr"]

            now = datetime.now()
            report_date = now.strftime("%d %b %Y")
            year = str(now.year)

            def _val(v, prefix="", suffix="", digits=0):
                if v is None:
                    return "—"
                try:
                    return f"{prefix}{_indian_format(v, digits)}{suffix}"
                except Exception:
                    return "—"

            target_str = _val(target)
            cmp_str = _val(current_price)
            mcap_str = _val(market_cap_cr)
            beta_str = _val(beta, digits=2) if beta is not None else "—"
            if upside is None:
                upside_str = "—"
            else:
                try:
                    u = float(upside)
                    upside_str = f"{'+' if u >= 0 else ''}{u:.0f}%"
                except Exception:
                    upside_str = "—"

            html = template_path.read_text(encoding="utf-8")

            # ── Specific, unique replacements ─────────────────────────────
            replacements = [
                ("[Report Type]", "INITIATING COVERAGE"),
                ("[Report Date]", report_date),
                ("[Analyst Name], [Credential]  ·  [email]",
                    "Research Team, CFA  ·  research@beaverintelligence.in"),
                ("[Company Name]", str(long_name)),
                ("[BLOOMBERG TICKER]", f"{ticker} IN EQUITY"),
                ("[NSE: TICKER]", f"NSE: {ticker}"),
                ("[Sector] · [Sub-sector]", f"{sector} · —"),
                ('<span class="rating-value">[BUY]</span>',
                    f'<span class="rating-value">{str(rating).upper()}</span>'),
                ('<span class="rating-tp">₹ [X,XXX]</span>',
                    f'<span class="rating-tp">₹ {target_str}</span>'),
                ('<span class="rating-upside">[+XX%]</span>',
                    f'<span class="rating-upside">{upside_str}</span>'),
                ('<span class="rating-meta">CMP ₹ [X,XXX] as of [Date]</span>',
                    f'<span class="rating-meta">CMP ₹ {cmp_str} as of {report_date}</span>'),
                ('<td class="value">₹ [X,XX,XXX] cr</td>',
                    f'<td class="value">₹ {mcap_str} cr</td>'),
                ('<td class="value">[X.XX]</td>', f'<td class="value">{beta_str}</td>'),
                ("[INH000000000]", "INH300009999"),
                ("[Year]", year),
                ("[YY]", "25"),
            ]

            # 52-week range, avg volume, free float from ctx (yfinance tier)
            wk_hi = ctx.get("week52_high")
            wk_lo = ctx.get("week52_low")
            if wk_hi is not None and wk_lo is not None:
                try:
                    range_str = f"₹ {_indian_format(wk_lo, 0)} — ₹ {_indian_format(wk_hi, 0)}"
                    replacements.append(('<td class="value">[X] — [Y]</td>',
                                         f'<td class="value">{range_str}</td>'))
                except Exception:
                    pass
            if ctx.get("avg_volume_cr") is not None:
                try:
                    replacements.append(('<td class="value">₹ [XX] cr</td>',
                                         f'<td class="value">₹ {_indian_format(ctx["avg_volume_cr"], 0)} cr</td>'))
                except Exception:
                    pass
            if ctx.get("free_float_pct") is not None:
                try:
                    replacements.append(('<td class="value">[XX]%</td>',
                                         f'<td class="value">{ctx["free_float_pct"]:.0f}%</td>'))
                except Exception:
                    pass

            # Peers
            peers = ctx.get("peers") or []
            for i, slot in enumerate(["[Peer 1]", "[Peer 2]", "[Peer 3]"]):
                replacements.append((slot, peers[i] if i < len(peers) else "—"))

            # Bull / bear TPs
            if ctx.get("bull_case_tp") is not None:
                try:
                    replacements.append(('<strong style="color: #0D3B2E;">₹ [X,XXX]</strong>',
                                         f'<strong style="color: #0D3B2E;">₹ {_indian_format(ctx["bull_case_tp"], 0)}</strong>'))
                except Exception:
                    pass
            if ctx.get("bear_case_tp") is not None:
                try:
                    replacements.append(('<strong style="color: #7A2028;">₹ [X,XXX]</strong>',
                                         f'<strong style="color: #7A2028;">₹ {_indian_format(ctx["bear_case_tp"], 0)}</strong>'))
                except Exception:
                    pass
            for old, new in replacements:
                html = html.replace(old, new)

            # Thesis one-liner
            if ctx.get("thesis_line"):
                html = re.sub(
                    r'(<div class="thesis">)[\s\S]*?(</div>)',
                    lambda m: m.group(1) + _html_escape(ctx["thesis_line"]) + m.group(2),
                    html, count=1,
                )

            # Executive summary paragraph (the 3-4 sentence bracketed block)
            if ctx.get("exec_summary"):
                html = re.sub(
                    r"\[Three to four sentences\.[^\]]*\]",
                    _html_escape(ctx["exec_summary"]),
                    html, count=1,
                )

            # Valuation-derivation caption
            if ctx.get("val_caption"):
                html = re.sub(
                    r"\[How the ₹\[X,XXX\] target price[^\]]*\]",
                    _html_escape(ctx["val_caption"]),
                    html, count=1,
                )

            # ── LLM-generated institutional commentary ────────────────────
            try:
                llm = generate_report_commentary(session_id, ticker, sector, ctx, session)
            except Exception as _e:
                logger.warning(f"[report-llm] wrapper failed: {_e}")
                llm = dict(_REPORT_EMPTY_CONTENT)

            def _rep_once(haystack: str, needle: str, value: str) -> str:
                if not value or needle not in haystack:
                    return haystack
                return haystack.replace(needle, _html_escape(value), 1)

            # Thesis (override if LLM gave one and tier1 didn't)
            if llm.get("thesis") and not ctx.get("thesis_line"):
                html = re.sub(
                    r'(<div class="thesis">)[\s\S]*?(</div>)',
                    lambda m: m.group(1) + _html_escape(llm["thesis"]) + m.group(2),
                    html, count=1,
                )

            # Exec summary — fill the standard bracketed prompt if still present
            if llm.get("exec_summary"):
                html = re.sub(
                    r"\[Three to four sentences\.[^\]]*\]",
                    _html_escape(llm["exec_summary"]),
                    html, count=1,
                )
            # Optional second paragraph — delete its placeholder
            html = re.sub(
                r"\[Optional second paragraph:[^\]]*\]", "", html, count=1,
            )

            # Rationale quadrants
            html = re.sub(r"\[One-line summary of consensus view[^\]]*\]",
                          _html_escape(llm.get("rationale_consensus") or "—"),
                          html, count=1)
            html = re.sub(r"\[Where you disagree and why[^\]]*\]",
                          _html_escape(llm.get("rationale_our_view")
                                       or ctx.get("rationale_our_view") or "—"),
                          html, count=1)
            html = re.sub(r"\[The specific catalyst that closes the gap[^\]]*\]",
                          _html_escape(llm.get("rationale_catalyst")
                                       or ctx.get("rationale_catalyst") or "—"),
                          html, count=1)
            html = re.sub(r"\[How the ₹\[X,XXX\] target price[^\]]*\]",
                          _html_escape(llm.get("rationale_valuation")
                                       or ctx.get("val_caption") or "—"),
                          html, count=1)

            # Named drivers / headwinds / opps / risks
            for slot, key in [
                ("[Driver 1]", "driver_1_name"),
                ("[Driver 2]", "driver_2_name"),
                ("[Driver 3]", "driver_3_name"),
                ("[Headwind 1]", "headwind_1_name"),
                ("[Headwind 2]", "headwind_2_name"),
                ("[Opportunity 1]", "opp_1_name"),
                ("[Opportunity 2]", "opp_2_name"),
                ("[Risk 1]", "risk_1_name"),
                ("[Risk 2]", "risk_2_name"),
            ]:
                html = _rep_once(html, slot, llm.get(key) or slot.strip("[]"))

            # Driver descriptions — repeat placeholder, replace in order
            driver_descs = [llm.get("driver_1_desc"), llm.get("driver_2_desc"),
                            llm.get("driver_3_desc")]
            for d in driver_descs:
                if not d:
                    continue
                html = html.replace("[One sentence. Specific, evidence-led, quantified.]",
                                    _html_escape(d), 1)
            # First driver may use the long examplified variant
            html = html.replace(
                '[One sentence. Specific, evidence-led, quantified where possible. '
                'E.g., "Studded share has risen 200 bps annually for four years; '
                'each 100 bps lifts jewellery EBIT margin by 40 bps."]',
                _html_escape(driver_descs[0] or "—"), 1,
            )

            # Headwind descriptions
            for d in [llm.get("headwind_1_desc"), llm.get("headwind_2_desc")]:
                if not d:
                    continue
                html = html.replace("[One sentence. The drag, quantified.]",
                                    _html_escape(d), 1)

            # Net view
            html = html.replace(
                "[One sentence explaining why the positives still outweigh — "
                "this is what the rating depends on.]",
                _html_escape(llm.get("net_view") or "—"), 1,
            )

            # Opportunity / risk descriptions
            opp_descs = [llm.get("opp_1_desc"), llm.get("opp_2_desc")]
            risk_descs = [llm.get("risk_1_desc"), llm.get("risk_2_desc")]
            # First opportunity uses the long example variant
            html = html.replace(
                '[Scenario · probability · TP impact. E.g., "If US expansion '
                'reaches 50 stores by FY30 vs. our 25, adds ₹300 to TP."]',
                _html_escape(opp_descs[0] or "—"), 1,
            )
            for d in opp_descs[1:] + risk_descs:
                if not d:
                    continue
                html = html.replace("[Scenario · probability · TP impact.]",
                                    _html_escape(d), 1)

            # Kill switch
            html = re.sub(
                r"\[A single, observable data point[\s\S]*?\]",
                _html_escape(llm.get("kill_switch") or "—"),
                html, count=1,
            )

            # ── Key Financials: revenue row + FY labels from DCF summary ──
            try:
                _dcf_sum = (session or {}).get("dcf_summary") or _load_dcf_summary(ticker)
                _fcast = (_dcf_sum or {}).get("forecast") or []
                logger.info(f"[report] forecast wiring: ticker={ticker} rows={len(_fcast)} source={'session' if (session or {}).get('dcf_summary') else 'disk'}")
                if _fcast:
                    years4 = _fcast[:4]
                    fy_labels = [f"FY{str(int(r['year']))[-2:]}E" for r in years4]
                    while len(fy_labels) < 4:
                        fy_labels.append("—")
                    # Whitespace-tolerant FY header replacement
                    html = re.sub(
                        r'<th class="num">FY23</th>\s*<th class="num">FY24</th>\s*'
                        r'<th class="num">FY25E</th>\s*<th class="num">FY26E</th>',
                        "\n              ".join(f'<th class="num">{l}</th>' for l in fy_labels),
                        html, count=1,
                    )

                    # Revenue row — convert millions → crores (÷10)
                    rev_cells = []
                    for i in range(4):
                        if i < len(years4) and years4[i].get("revenue") is not None:
                            try:
                                rev_cr = int(float(years4[i]["revenue"]) / 10)
                                rev_cells.append(f'<td class="num">{rev_cr:,}</td>')
                            except Exception:
                                rev_cells.append('<td class="num">—</td>')
                        else:
                            rev_cells.append('<td class="num">—</td>')
                    rev_row_new = (
                        '<tr><td class="label-cell">Revenue</td>'
                        + "".join(rev_cells) + "</tr>"
                    )
                    # Whitespace-tolerant Revenue row replacement
                    html, _n = re.subn(
                        r'<tr>\s*<td class="label-cell">Revenue</td>\s*'
                        r'(?:<td class="num">\[X,XXX\]</td>\s*){4}</tr>',
                        rev_row_new.replace("\\", r"\\"),
                        html, count=1,
                    )
                    logger.info(f"[report] wired {len(years4)}y forecast revenue row for {ticker} (matches={_n})")
            except Exception as _e:
                logger.warning(f"[report] forecast wiring failed: {_e}")

            # ── Belt-and-suspenders: safe rating read straight from disk ──
            try:
                _sess_dir = f"/app/backend/research_platform/ai_engine/sessions/{session_id}"
                with open(f"{_sess_dir}/scenarios.json") as _f:
                    _scen = json.load(_f)
                _r = (_scen.get("scenarios", {}).get("base", {}).get("rating")
                      or _scen.get("base", {}).get("rating")
                      or "HOLD")
                rating = str(_r).upper()
            except Exception as _e:
                logger.debug(f"[report] disk rating read fallback: {_e}")
            if not rating or str(rating).strip() in ("", "—"):
                rating = "HOLD"
            html = html.replace("[BUY]", str(rating).upper())

            # ── Final cleanup pass ────────────────────────────────────────
            short_labels = [
                "[Driver 1]", "[Driver 2]", "[Driver 3]",
                "[Headwind 1]", "[Headwind 2]",
                "[Opportunity 1]", "[Opportunity 2]",
                "[Risk 1]", "[Risk 2]",
                "[Peer 1]", "[Peer 2]", "[Peer 3]",
                "[Subject]",
                "[X,XX,XXX]", "[X,XXX]", "[XXX]", "[XX.X]", "[XX]",
                "[+XX]", "[XXx]", "[X.X]", "[X]", "[Y]", "[Date]",
                "[Name, Credential]", "[Analyst Name, Credential]",
                "[INH000000000]", "[Year]",
                "[Report Type]", "[Report Date]", "[email]",
                "[Company Name]", "[Ticker]", "[BLOOMBERG TICKER]",
                "[Sub-sector]",
            ]
            for label in short_labels:
                html = html.replace(label, "—")

            html = html.replace("[XX.X]%", "—")
            html = html.replace("[XX]%", "—")
            html = html.replace("[+XX]%", "—")
            html = html.replace("[+XX] bps", "—")
            html = html.replace("₹ [X,XXX]", "₹ —")
            html = html.replace("₹[X,XXX]", "₹—")
            html = html.replace("₹ [X,XX,XXX]", "₹ —")
            html = html.replace("₹[XX]", "₹—")
            html = html.replace("FY[YY]E", "FY25E")
            html = html.replace("FY[YY]", "FY25")

            # Strip long instructional brackets (sentences)
            html = re.sub(r"\[[^\]]{20,}\]", "—", html)
            # Final sweep: short word-only brackets still remaining
            html = re.sub(r"\[[A-Za-z][a-zA-Z ]{1,25}\d?\]", "—", html)

            filename = f"BEAVER_{ticker}_research_report.html"
            return Response(
                content=html,
                media_type="text/html; charset=utf-8",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                },
            )
    except Exception as e:
        logger.warning(f"[report] template render failed, falling back: {e}")

    # ─── Fallback: research_platform pdf_builder ─────────────────────────
    if RP_AVAILABLE and RP_PDF and rp_build_report:
        try:
            # Find the matching filesystem session — try by ID first, then by ticker
            rp_ses = _find_rp_session(session_id)
            if rp_ses is None:
                # Look up ticker from Mongo and find latest RP session for it
                db = _get_db()
                if db is not None:
                    try:
                        doc = await db.sessions.find_one({"_id": session_id})
                        if doc and doc.get("ticker"):
                            rp_ses = _find_rp_session_by_ticker(doc["ticker"])
                    except Exception:
                        pass
            if rp_ses is not None:
                # Determine sector for report styling
                meta = rp_ses.get_meta() if hasattr(rp_ses, "get_meta") else {}
                sector = meta.get("sector") or meta.get("_sector") or \
                         _detect_sector_simple(rp_ses.ticker)
                # Generate the real 2-page A4 report
                html_path = rp_build_report(rp_ses, sector=sector)
                html_content = Path(html_path).read_text(encoding="utf-8")
                logger.info(f"[report] served RP report for {session_id} → {rp_ses.session_id}")
                return HTMLResponse(content=html_content, status_code=200)
        except Exception as e:
            logger.warning(f"[report] RP pdf_builder failed, falling back: {e}")

    # ─── Fallback: simple HTML report from cached data ────────────────────
    try:
        session = await _load_session_for_report(session_id)
        ticker = session.get("ticker", "UNKNOWN")
        price = session.get("price") or session.get("current_price") or (session.get("dcf") or {}).get("current_price")
        prev_close = session.get("prev_close")
        change_pct = session.get("change_percent") or session.get("change_pct")
        dcf = session.get("dcf") or {}
        hypothesis = session.get("hypothesis", "—")
        sector = session.get("sector", "general")
        created = session.get("created_at", "")
        status = session.get("status", "active")

        def fmt(v, suffix=""):
            if v is None:
                return "—"
            try:
                return f"{float(v):,.2f}{suffix}"
            except Exception:
                return str(v)

        html = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8"><title>{ticker} Research Report</title>
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; max-width: 900px; margin: 2rem auto; padding: 2rem; color: #0f172a; }}
  h1 {{ border-bottom: 2px solid #2563eb; padding-bottom: 0.5rem; margin-bottom: 0.5rem; }}
  h2 {{ color: #334155; margin-top: 2rem; border-bottom: 1px solid #e5e7eb; padding-bottom: 0.3rem; }}
  .meta {{ color: #64748b; font-size: 0.9rem; margin-bottom: 2rem; }}
  .card {{ background: #f8fafc; border: 1px solid #e5e7eb; border-radius: 8px; padding: 1rem; margin: 1rem 0; }}
  .kv {{ display: grid; grid-template-columns: 200px 1fr; gap: 0.5rem; margin: 0.25rem 0; }}
  .kv .k {{ color: #64748b; font-weight: 500; }}
  .kv .v {{ color: #0f172a; font-weight: 600; font-variant-numeric: tabular-nums; }}
  .positive {{ color: #16a34a; }}
  .negative {{ color: #dc2626; }}
  .footer {{ margin-top: 3rem; padding-top: 1rem; border-top: 1px solid #e5e7eb; color: #94a3b8; font-size: 0.8rem; }}
</style></head>
<body>
  <h1>{ticker}</h1>
  <div class="meta">
    Sector: {sector} &middot; Session: {session_id} &middot; Status: {status}<br>
    Generated: {datetime.now(timezone.utc).isoformat()} &middot; Created: {created}
  </div>

  <h2>Price Snapshot</h2>
  <div class="card">
    <div class="kv"><span class="k">Current Price</span><span class="v">₹{fmt(price)}</span></div>
    <div class="kv"><span class="k">Previous Close</span><span class="v">₹{fmt(prev_close)}</span></div>
    <div class="kv"><span class="k">Change (%)</span><span class="v {'positive' if (change_pct or 0) >= 0 else 'negative'}">{fmt(change_pct, '%')}</span></div>
  </div>

  <h2>DCF Valuation</h2>
  <div class="card">
    <div class="kv"><span class="k">Fair Value (Base)</span><span class="v">₹{fmt(dcf.get('fair_value'))}</span></div>
    <div class="kv"><span class="k">Upside vs Current</span><span class="v {'positive' if (dcf.get('upside_pct') or 0) >= 0 else 'negative'}">{fmt(dcf.get('upside_pct'), '%')}</span></div>
    <div class="kv"><span class="k">WACC</span><span class="v">{fmt((dcf.get('wacc') or 0) * 100, '%')}</span></div>
    <div class="kv"><span class="k">Growth Rate</span><span class="v">{fmt((dcf.get('growth') or 0) * 100, '%')}</span></div>
    <div class="kv"><span class="k">Terminal Growth</span><span class="v">{fmt((dcf.get('terminal_growth') or 0) * 100, '%')}</span></div>
    <div class="kv"><span class="k">Enterprise Value</span><span class="v">₹{fmt(dcf.get('enterprise_value'))}</span></div>
    <div class="kv"><span class="k">Equity Value</span><span class="v">₹{fmt(dcf.get('equity_value'))}</span></div>
  </div>

  <h2>Hypothesis</h2>
  <div class="card"><p>{hypothesis}</p></div>

  <div class="footer">
    Market Pulse Research · Automated report · Figures from yfinance at time of analysis.
    Not investment advice. Verify all numbers independently before making decisions.
  </div>
</body></html>"""
        return HTMLResponse(content=html, status_code=200)
    except Exception as e:
        logger.error(f"report_download failed: {e}\n{traceback.format_exc()}")
        return HTMLResponse(
            content=f"<html><body><h1>Report unavailable</h1><p>{str(e)}</p></body></html>",
            status_code=200,
        )


@api_router.get("/research/{session_id}/report/xlsx")
async def report_download_xlsx(session_id: str):
    """
    Return XLSX spreadsheet with session details.
    Prefers research_platform's 6-sheet Excel (assumptions, history, scenarios,
    sensitivity, insights, sources) if available.
    Falls back to a simple one-sheet summary if that fails.
    """
    from fastapi.responses import Response

    # ─── Try research_platform's audit_export first ───────────────────────
    if RP_AVAILABLE and RP_EXPORT and rp_export_excel:
        try:
            rp_ses = _find_rp_session(session_id)
            if rp_ses is None:
                db = _get_db()
                if db is not None:
                    try:
                        doc = await db.sessions.find_one({"_id": session_id})
                        if doc and doc.get("ticker"):
                            rp_ses = _find_rp_session_by_ticker(doc["ticker"])
                    except Exception:
                        pass
            if rp_ses is not None:
                xlsx_path = rp_export_excel(rp_ses)
                content = Path(xlsx_path).read_bytes()
                ticker = rp_ses.ticker
                logger.info(f"[report] served RP xlsx for {session_id} → {rp_ses.session_id} ({len(content)} bytes)")
                return Response(
                    content=content,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{ticker}_research.xlsx"'},
                )
        except Exception as e:
            logger.warning(f"[report] RP audit_export failed, falling back: {e}")

    # ─── Fallback: simple one-sheet XLSX ──────────────────────────────────
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            # openpyxl not installed → fall back to CSV
            return await report_download_csv(session_id)

        session = await _load_session_for_report(session_id)
        ticker = session.get("ticker", "UNKNOWN")
        price = session.get("price") or (session.get("dcf") or {}).get("current_price")
        prev_close = session.get("prev_close")
        change_pct = session.get("change_percent") or session.get("change_pct")
        dcf = session.get("dcf") or {}

        wb = Workbook()

        # Sheet 1: Summary
        ws = wb.active
        ws.title = "Summary"
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")
        section_font = Font(bold=True, size=11, color="1E40AF")
        label_font = Font(color="64748B")

        ws['A1'] = f"{ticker} — Research Report"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:B1')
        ws.row_dimensions[1].height = 28

        rows = [
            ("", ""),
            ("SESSION DETAILS", ""),
            ("Ticker", ticker),
            ("Sector", session.get("sector", "general")),
            ("Session ID", session_id),
            ("Status", session.get("status", "active")),
            ("Created", session.get("created_at", "")),
            ("Generated", datetime.now(timezone.utc).isoformat()),
            ("", ""),
            ("PRICE", ""),
            ("Current Price", price),
            ("Previous Close", prev_close),
            ("Change %", f"{change_pct:.2f}%" if change_pct is not None else "—"),
            ("", ""),
            ("DCF VALUATION", ""),
            ("Fair Value (Base)", dcf.get("fair_value")),
            ("Upside %", f"{dcf.get('upside_pct'):.2f}%" if dcf.get("upside_pct") is not None else "—"),
            ("WACC", f"{dcf.get('wacc') * 100:.2f}%" if dcf.get("wacc") else "—"),
            ("Growth Rate", f"{dcf.get('growth') * 100:.2f}%" if dcf.get("growth") else "—"),
            ("Terminal Growth", f"{dcf.get('terminal_growth') * 100:.2f}%" if dcf.get("terminal_growth") else "—"),
            ("Enterprise Value", dcf.get("enterprise_value")),
            ("Equity Value", dcf.get("equity_value")),
            ("", ""),
            ("HYPOTHESIS", ""),
            ("Thesis", session.get("hypothesis", "—")),
        ]
        for i, (k, v) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v)
            if k in ("SESSION DETAILS", "PRICE", "DCF VALUATION", "HYPOTHESIS"):
                ws.cell(row=i, column=1).font = section_font
            elif k:
                ws.cell(row=i, column=1).font = label_font

        # Save to bytes
        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{ticker}_report.xlsx"'},
        )
    except Exception as e:
        logger.error(f"report_xlsx failed: {e}\n{traceback.format_exc()}")
        # Fallback to CSV on any error
        return await report_download_csv(session_id)


@api_router.get("/research/{session_id}/report/csv")
async def report_download_csv(session_id: str):
    """Return a CSV version of the report. Always works, no dependencies."""
    from fastapi.responses import Response
    try:
        session = await _load_session_for_report(session_id)
        ticker = session.get("ticker", "UNKNOWN")
        dcf = session.get("dcf") or {}
        rows = [
            ("Ticker", ticker),
            ("Sector", session.get("sector", "general")),
            ("Session ID", session_id),
            ("Status", session.get("status", "active")),
            ("Created", session.get("created_at", "")),
            ("Current Price", session.get("price") or dcf.get("current_price") or ""),
            ("Previous Close", session.get("prev_close") or ""),
            ("Change %", session.get("change_percent") or session.get("change_pct") or ""),
            ("Fair Value", dcf.get("fair_value") or ""),
            ("Upside %", dcf.get("upside_pct") or ""),
            ("WACC", dcf.get("wacc") or ""),
            ("Growth", dcf.get("growth") or ""),
            ("Terminal Growth", dcf.get("terminal_growth") or ""),
            ("Enterprise Value", dcf.get("enterprise_value") or ""),
            ("Equity Value", dcf.get("equity_value") or ""),
            ("Hypothesis", session.get("hypothesis", "")),
        ]
        csv_lines = ["Field,Value"] + [f'"{k}","{v}"' for k, v in rows]
        csv_content = "\n".join(csv_lines)
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{ticker}_report.csv"'},
        )
    except Exception as e:
        logger.error(f"report_csv failed: {e}")
        return Response(content="Field,Value\nerror," + str(e), media_type="text/csv", status_code=200)


@api_router.get("/research/{session_id}/report/markdown")
async def report_download_markdown(session_id: str):
    """Return summary.md as a download attachment."""
    from fastapi.responses import Response
    try:
        session = await _load_session_for_report(session_id)
        ticker = (session.get("ticker") or "report").upper()
        rp_ses = _find_rp_session(session_id)
        sdir = None
        if rp_ses is not None:
            sdir = (getattr(rp_ses, "session_dir", None)
                    or getattr(rp_ses, "path", None)
                    or getattr(rp_ses, "dir", None))
        if sdir:
            sfp = Path(sdir) / "summary.md"
            if sfp.exists():
                content = sfp.read_text(encoding="utf-8")
                return Response(
                    content=content.encode("utf-8"),
                    media_type="text/markdown; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="{ticker}_summary.md"'},
                )
        # Fallback: generate minimal markdown from session data so the download always works
        thesis = session.get("thesis") or session.get("hypothesis") or "Analysis pending."
        scoring = session.get("scoring") or {}
        rec = scoring.get("recommendation", "N/A")
        score = scoring.get("composite_score", "N/A")
        fallback_md = (
            f"# {ticker} — Research Summary\n\n"
            f"**Recommendation:** {rec}  \n"
            f"**Composite Score:** {score}  \n\n"
            f"## Investment Thesis\n\n{thesis}\n\n"
            f"*Full markdown summary available after DCF analysis completes.*\n"
        )
        return Response(
            content=fallback_md.encode("utf-8"),
            media_type="text/markdown; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{ticker}_summary.md"'},
        )
    except Exception as e:
        logger.error(f"report_markdown failed: {e}")
        return Response(
            content=f"# Error\n\n{e}".encode("utf-8"),
            media_type="text/markdown; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="error.md"'},
        )


@api_router.get("/research/{session_id}/audit")
async def audit_trail(session_id: str):
    rp_ses = _find_rp_session(session_id)
    sources = []
    audit_log = []
    if rp_ses is not None:
        # sources.json if it exists
        raw_sources = _read_rp_session_file(rp_ses, "sources.json")
        if isinstance(raw_sources, list):
            sources = raw_sources
        elif isinstance(raw_sources, dict):
            sources = raw_sources.get("sources", [])
        # Fall back to session_meta for data source info
        if not sources:
            meta = _read_rp_session_file(rp_ses, "session_meta.json")
            asm = _read_rp_session_file(rp_ses, "assumptions.json")
            data_src = asm.get("_data_source", meta.get("data_source", "yfinance"))
            sources = [{"api_name": data_src, "endpoint": "yfinance.Ticker", "fetched_at": meta.get("created_at", ""), "status": "ok"}]
        # audit_log.json if it exists, else assumptions_history
        raw_audit = _read_rp_session_file(rp_ses, "audit_log.json")
        if isinstance(raw_audit, list):
            audit_log = raw_audit
        else:
            hist = _read_rp_session_file(rp_ses, "assumptions_history.json")
            if isinstance(hist, list):
                for entry in hist:
                    evt = entry.get("event", "update")
                    ts = (entry.get("assumptions") or {}).get("_initialized_at", "")
                    audit_log.append({"timestamp": ts, "event": evt, "fields_changed": list(entry.get("delta", {}).keys())})
    return {"sources": sources, "audit_log": audit_log}


@api_router.get("/research/{session_id}/guardrails")
async def guardrail_log(session_id: str):
    rp_ses = _find_rp_session(session_id)
    raw = _read_rp_session_file(rp_ses, "guardrail_log.json")
    breaches = raw if isinstance(raw, list) else []
    return {"guardrails": breaches, "all_passed": len(breaches) == 0}


@api_router.get("/research/{session_id}/assumption_history")
async def assumption_history(session_id: str):
    rp_ses = _find_rp_session(session_id)
    raw = _read_rp_session_file(rp_ses, "assumptions_history.json")
    history = raw if isinstance(raw, list) else []
    return {"history": history}


@api_router.get("/macro")
@safe_endpoint(lambda: {"indicators": [], "globalEvents": [], "macroMicro": []})
async def macro():
    indicators = []
    global_events = []
    try:
        from sqlalchemy import select, desc
        from research_platform.database.connection import get_session as _rp_dbs
        from research_platform.database.models import MacroIndicator as _MI, FxRate as _FX, Event as _Ev
        with _rp_dbs() as db:
            mi_rows = db.scalars(select(_MI).order_by(desc(_MI.date)).limit(200)).all()
            for r in mi_rows:
                indicators.append({
                    "id": r.indicator,
                    "title": r.indicator,
                    "value": f"{float(r.value):.2f}" if r.value is not None else "—",
                    "raw_value": float(r.value) if r.value is not None else None,
                    "changeType": "neutral",
                    "change": "—",
                    "subtitle": f"{r.source or 'FRED'} · {r.date}",
                })
            # Fall back to FX rates when macro_indicators table is empty
            if not indicators:
                fx_rows = db.scalars(select(_FX).order_by(desc(_FX.date)).limit(20)).all()
                seen_pairs: set = set()
                for r in fx_rows:
                    if r.pair not in seen_pairs:
                        seen_pairs.add(r.pair)
                        indicators.append({
                            "id": r.pair,
                            "title": r.pair,
                            "value": f"{float(r.rate):.4f}",
                            "raw_value": float(r.rate),
                            "changeType": "neutral",
                            "change": "—",
                            "subtitle": f"{r.source or 'Frankfurter'} · {r.date}",
                        })
            ev_rows = db.scalars(
                select(_Ev).where(_Ev.type == "regulatory").order_by(desc(_Ev.date)).limit(8)
            ).all()
            for r in ev_rows:
                global_events.append({
                    "id": r.id,
                    "event": r.title,
                    "region": "India",
                    "impact": "Neutral",
                })
    except Exception as exc:
        logger.debug(f"[macro] DB read failed: {exc}")
    return {"indicators": indicators, "globalEvents": global_events, "macroMicro": []}

@api_router.get("/signals")
@safe_endpoint(lambda: {"signals": []})
async def signals():
    signals_out = []
    try:
        from sqlalchemy import select, desc
        from research_platform.database.connection import get_session as _rp_dbs
        from research_platform.database.models import Event as _Ev
        with _rp_dbs() as db:
            rows = db.scalars(
                select(_Ev)
                .where(_Ev.type.in_(["signal", "regulatory"]))
                .order_by(desc(_Ev.date))
                .limit(100)
            ).all()
            for r in rows:
                et = r.entity_type or ""
                parts = et.split(":") if et else []
                sector = parts[0] if parts else ""
                direction = parts[1] if len(parts) > 1 else "neutral"
                score = r.impact_score or 0.0
                if direction == "positive":
                    sev = "positive"
                elif direction == "negative":
                    sev = "danger" if score >= 1.0 else "negative"
                else:
                    sev = "warning" if score >= 1.0 else "info"
                signals_out.append({
                    "id": r.id,
                    "title": r.title,
                    "timestamp": str(r.date) if r.date else "",
                    "severity": sev,
                    "sector": sector,
                    "signalType": r.type.capitalize(),
                    "source_url": r.source_url or "",
                })
    except Exception as exc:
        logger.debug(f"[signals] DB read failed: {exc}")
    return {"signals": signals_out}

@api_router.get("/alerts")
@safe_endpoint(lambda: {"alerts": []})
async def alerts():
    return {"alerts": []}

_SECTOR_DISPLAY = {
    "petroleum_energy": "Petroleum & Energy",
    "banking_nbfc":     "Banking & NBFC",
    "it_tech":          "IT & Technology",
    "pharma":           "Pharma & Healthcare",
    "fmcg_retail":      "FMCG & Retail",
    "real_estate":      "Real Estate",
    "auto":             "Automobile",
    "other":            "Universal",
}

_SECTOR_VALUATION_FOCUS = {
    "petroleum_energy": "EV/EBITDA + DCF",
    "banking_nbfc":     "P/B + P/E + DDM",
    "it_tech":          "P/E + DCF + EV/EBITDA",
    "pharma":           "P/E + EV/EBITDA",
    "fmcg_retail":      "P/E + DCF",
    "real_estate":      "NAV + EV/EBITDA",
    "auto":             "EV/EBITDA + P/E",
    "other":            "DCF + Comps",
}

_SECTOR_DIR_MAP = {
    "petroleum_energy": "petroleum",
    "banking_nbfc":     "banking",
    "it_tech":          "it",
    "pharma":           "pharma",
    "fmcg_retail":      "fmcg",
    "real_estate":      "real_estate",
    "auto":             "auto",
}

@api_router.get("/research/sector/{sector_name}")
async def sector_framework(sector_name: str):
    dir_name = _SECTOR_DIR_MAP.get(sector_name, sector_name)
    sector_file = Path("/app/backend/research_platform/ai_engine/frameworks") / dir_name / "signals.json"
    key_drivers = []
    key_metrics = []
    if sector_file.exists():
        try:
            import json as _json
            data = _json.loads(sector_file.read_text())
            key_drivers = data.get("key_drivers", [])
            key_metrics = data.get("key_metrics", data.get("metrics", []))
        except Exception:
            pass
    return {
        "name": sector_name,
        "display_name": _SECTOR_DISPLAY.get(sector_name, sector_name.replace("_", " ").title()),
        "key_drivers": key_drivers[:6],
        "key_metrics": key_metrics[:6],
        "valuation_focus": _SECTOR_VALUATION_FOCUS.get(sector_name, "DCF + Comps"),
    }

@api_router.get("/derivatives/{ticker}")
async def derivatives(ticker: str):
    """Return live F&O data for a ticker from NSE option chain."""
    import httpx
    ticker_up = ticker.strip().upper()
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Accept": "application/json",
            "Referer": "https://www.nseindia.com/",
            "Accept-Language": "en-US,en;q=0.9",
        }
        async with httpx.AsyncClient(timeout=15, follow_redirects=True) as client:
            await client.get("https://www.nseindia.com/", headers=headers)
            url = f"https://www.nseindia.com/api/option-chain-equities?symbol={ticker_up}"
            resp = await client.get(url, headers=headers)
            resp.raise_for_status()
            data = resp.json()

        filtered = data.get("filtered") or {}
        chain = filtered.get("data") or []

        total_ce_oi = sum(d.get("CE", {}).get("openInterest", 0) for d in chain if d.get("CE"))
        total_pe_oi = sum(d.get("PE", {}).get("openInterest", 0) for d in chain if d.get("PE"))
        total_ce_vol = sum(d.get("CE", {}).get("totalTradedVolume", 0) for d in chain if d.get("CE"))
        total_pe_vol = sum(d.get("PE", {}).get("totalTradedVolume", 0) for d in chain if d.get("PE"))

        pcr_oi = round(total_pe_oi / total_ce_oi, 2) if total_ce_oi > 0 else None
        pcr_vol = round(total_pe_vol / total_ce_vol, 2) if total_ce_vol > 0 else None
        total_oi = total_ce_oi + total_pe_oi

        # Max OI strike
        max_oi_strike = None
        max_oi_val = 0
        for d in chain:
            for side in ("CE", "PE"):
                if d.get(side) and d[side].get("openInterest", 0) > max_oi_val:
                    max_oi_val = d[side]["openInterest"]
                    max_oi_strike = d.get("strikePrice")

        oi_change_pct = None
        try:
            ce_oi_chg = sum(d["CE"].get("changeinOpenInterest", 0) for d in chain if d.get("CE"))
            pe_oi_chg = sum(d["PE"].get("changeinOpenInterest", 0) for d in chain if d.get("PE"))
            base_oi = total_oi - (ce_oi_chg + pe_oi_chg)
            if base_oi > 0:
                oi_change_pct = round((ce_oi_chg + pe_oi_chg) / base_oi * 100, 1)
        except Exception:
            pass

        sentiment = (
            "bullish" if pcr_oi and pcr_oi > 1.2 else
            "bearish" if pcr_oi and pcr_oi < 0.7 else
            "neutral" if pcr_oi else None
        )

        return {
            "pcr_oi": pcr_oi,
            "pcr_volume": pcr_vol,
            "max_oi_strike": max_oi_strike,
            "total_oi": total_oi,
            "oi_change_pct": oi_change_pct,
            "sentiment": sentiment,
        }
    except Exception as e:
        logger.debug(f"[derivatives] {ticker_up} NSE failed ({e}), trying yfinance fallback")

    # yfinance fallback — options chain for Indian stocks via .NS suffix
    if YF_AVAILABLE:
        try:
            import asyncio as _asyncio
            loop = _asyncio.get_event_loop()
            def _yf_options():
                ns_ticker = ticker_up if ticker_up.endswith(".NS") else f"{ticker_up}.NS"
                tk = yf.Ticker(ns_ticker)
                expirations = tk.options
                if not expirations:
                    return None
                chain = tk.option_chain(expirations[0])
                calls = chain.calls
                puts  = chain.puts
                total_call_oi = int(calls["openInterest"].fillna(0).sum())
                total_put_oi  = int(puts["openInterest"].fillna(0).sum())
                total_call_vol = int(calls["volume"].fillna(0).sum())
                total_put_vol  = int(puts["volume"].fillna(0).sum())
                pcr_oi  = round(total_put_oi / total_call_oi, 2) if total_call_oi else None
                pcr_vol = round(total_put_vol / total_call_vol, 2) if total_call_vol else None
                # Max OI strike from combined chain
                all_oi = (
                    calls[["strike", "openInterest"]].rename(columns={"openInterest": "oi"})
                    .append(puts[["strike", "openInterest"]].rename(columns={"openInterest": "oi"}),
                            ignore_index=True)
                )
                max_oi_strike = None
                if not all_oi.empty:
                    max_oi_strike = float(all_oi.loc[all_oi["oi"].idxmax(), "strike"])
                sentiment = (
                    "bullish" if pcr_oi and pcr_oi > 1.2 else
                    "bearish" if pcr_oi and pcr_oi < 0.7 else
                    "neutral" if pcr_oi else None
                )
                return {
                    "pcr_oi": pcr_oi,
                    "pcr_volume": pcr_vol,
                    "max_oi_strike": max_oi_strike,
                    "total_oi": total_call_oi + total_put_oi,
                    "oi_change_pct": None,
                    "sentiment": sentiment,
                    "expiry": expirations[0],
                    "source": "yfinance",
                }
            result = await loop.run_in_executor(None, _yf_options)
            if result:
                logger.debug(f"[derivatives] {ticker_up} yfinance fallback OK PCR={result.get('pcr_oi')}")
                return result
        except Exception as _yfe:
            logger.debug(f"[derivatives] {ticker_up} yfinance fallback failed: {_yfe}")

    return {
        "pcr_oi": None, "pcr_volume": None, "max_oi_strike": None,
        "total_oi": None, "oi_change_pct": None, "sentiment": None,
    }


@api_router.get("/insider-trades")
async def insider_trades(ticker: str = "", limit: int = 20):
    """Return insider/bulk deals for a ticker from the research_platform event table."""
    from datetime import date as _date, timedelta as _td
    results = []
    try:
        from database.connection import get_session as _get_rp_session
        from database.models import Event
        from sqlalchemy import select as _select, and_ as _and, func as _func
        ticker_upper = ticker.strip().upper()
        cutoff = _date.today() - _td(days=90)
        with _get_rp_session() as s:
            stmt = (
                _select(Event)
                .where(
                    _and(
                        Event.entity_type == "insider_trade",
                        Event.date >= cutoff,
                    )
                )
                .order_by(Event.date.desc())
                .limit(200)
            )
            rows = s.execute(stmt).scalars().all()
        for ev in rows:
            title = ev.title or ""
            # Title format: "Bulk Deal: TICKER — Person BUY qty @ ₹price (₹xCr)"
            # or "BSE Bulk Deal: TICKER — ..."
            if ticker_upper and ticker_upper not in title.upper():
                continue
            # Parse person/action/qty from title
            person, action, qty_str, value_inr = "—", "—", "—", None
            trade_type = "BULK"
            try:
                import re as _re
                m = _re.search(r"(?:Bulk Deal|BSE Bulk Deal):\s*\S+\s*[—–-]\s*(.+?)\s+(BUY|SELL|B|S)\s+([\d,]+)\s*@\s*[₹]?([\d.]+)", title, _re.I)
                if m:
                    person = m.group(1).strip()
                    raw_act = m.group(2).upper()
                    action = "BUY" if raw_act in ("BUY", "B") else "SELL"
                    qty_str = m.group(3).replace(",", "")
                    price_str = m.group(4)
                    try:
                        qty_i = int(qty_str)
                        price_f = float(price_str)
                        value_inr = round(qty_i * price_f)
                    except Exception:
                        pass
                else:
                    parts = title.split("—", 1)
                    if len(parts) > 1:
                        rest = parts[1].strip()
                        person = rest[:80]
            except Exception:
                pass
            results.append({
                "date": str(ev.date) if ev.date else "",
                "person": person,
                "action": action,
                "quantity": qty_str,
                "value_inr": value_inr,
                "type": trade_type,
            })
            if len(results) >= limit:
                break
    except Exception as e:
        logger.debug(f"[insider-trades] DB unavailable: {e}")
    return {"trades": results}


@api_router.get("/news")
async def news_feed(limit: int = 10):
    try:
        items = await fetch_news_safe()
        return {"news": items[:limit]}
    except Exception:
        return {"news": []}

@api_router.get("/prices/{ticker}")
@safe_endpoint(lambda: {"price": None, "error": "unavailable"})
async def prices(ticker: str):
    ticker_ns = ticker if "." in ticker else f"{ticker}.NS"
    if not YF_AVAILABLE:
        return {"price": None, "ticker": ticker_ns, "source": "unavailable"}
    try:
        tk = yf.Ticker(ticker_ns)
        hist = tk.history(period="1d")
        if hist is not None and not hist.empty:
            return {
                "ticker": ticker_ns,
                "price": _safe_float(hist["Close"].iloc[-1]),
                "source": "yfinance",
            }
    except Exception as e:
        logger.warning(f"prices({ticker}) failed: {e}")
    return {"price": None, "ticker": ticker_ns, "source": "error"}

@api_router.post("/chat")
@safe_endpoint(lambda: {"reply": "Chat service unavailable right now. Please try again in a moment."})
async def chat(request: Request):
    body = await request.json()
    message = body.get("message", "").strip()
    if not message:
        return {"reply": "Please provide a message."}
    if not ANTHROPIC_KEY:
        return {"reply": "AI chat is not configured on this deployment."}
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
        resp = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=512,
            messages=[{"role": "user", "content": message}],
        )
        text = "".join(b.text for b in resp.content if hasattr(b, "text"))
        return {"reply": text or "No response."}
    except Exception as e:
        logger.warning(f"chat failed: {e}")
        return {"reply": "Chat temporarily unavailable."}


BEAVER_CHAT_SYSTEM_PROMPT = (
    "You are Beaver AI, an equity research assistant focused on Indian "
    "markets (NSE and BSE). You help users understand stocks, DCF valuation, "
    "scenarios, and macro factors. Keep responses concise and professional "
    "— 2-3 short paragraphs max. If asked about a specific stock, give "
    "genuinely useful analytical context. If asked something off-topic, "
    "politely redirect to equity research questions."
)


@api_router.post("/chat/message")
async def chat_message(request: Request):
    """Beaver AI chat — Claude Sonnet, Indian-equities-focused assistant."""
    from fastapi.responses import JSONResponse
    try:
        body = await request.json()
    except Exception:
        body = {}
    message = (body.get("message") or "").strip()
    if not message:
        return JSONResponse(status_code=400, content={"error": "message is required"})
    if not ANTHROPIC_KEY:
        return JSONResponse(
            status_code=503,
            content={"error": "Chat service not configured"},
        )
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
        resp = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=500,
            system=BEAVER_CHAT_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": message}],
        )
        text = "".join(b.text for b in resp.content if hasattr(b, "text"))
        return {"response": text or "No response."}
    except Exception as e:
        logger.error(f"chat_message failed: {e}\n{traceback.format_exc()}")
        return JSONResponse(
            status_code=500,
            content={"error": "Couldn't get a response, try again"},
        )


# ─────────────────────────────────────────────────────────────────────────────
# FEATURE 14 — Financial Charts
# ─────────────────────────────────────────────────────────────────────────────

def _generate_charts(ticker: str) -> list[dict]:
    """Generate 6 financial charts as base64 PNGs using yfinance + matplotlib."""
    import io, base64
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    import yfinance as yf
    import numpy as np

    charts = []

    def _b64(fig) -> str:
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=90, bbox_inches="tight", facecolor=fig.get_facecolor())
        buf.seek(0)
        plt.close(fig)
        return base64.b64encode(buf.read()).decode()

    ticker_yf = ticker if (ticker.endswith(".NS") or ticker.endswith(".BO")) else ticker + ".NS"
    t = yf.Ticker(ticker_yf)

    # ── Chart 1: Price History (1Y) ──────────────────────────────────────────
    try:
        hist = t.history(period="1y")
        if not hist.empty:
            fig, ax = plt.subplots(figsize=(6, 3), facecolor="#0f172a")
            ax.set_facecolor("#1e293b")
            ax.plot(hist.index, hist["Close"], color="#22d3ee", linewidth=1.5)
            ax.fill_between(hist.index, hist["Close"], alpha=0.15, color="#22d3ee")
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%b '%y"))
            ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
            ax.tick_params(colors="#94a3b8", labelsize=8)
            ax.spines[:].set_color("#334155")
            ax.yaxis.tick_right()
            ax.set_title("Price History (1Y)", color="#e2e8f0", fontsize=9, pad=6)
            plt.setp(ax.get_xticklabels(), rotation=30, ha="right")
            charts.append({"name": "Price History (1Y)", "type": "line", "image_base64": _b64(fig)})
    except Exception as e:
        logger.debug(f"[charts] price history failed: {e}")

    # ── Chart 2: Revenue (3Y) ────────────────────────────────────────────────
    try:
        fin = t.financials
        if fin is not None and not fin.empty and "Total Revenue" in fin.index:
            rev_row = fin.loc["Total Revenue"].dropna()
            years = [str(c)[:4] for c in rev_row.index[:4]][::-1]
            vals = [float(v) / 1e7 for v in rev_row.values[:4]][::-1]  # ₹ Cr
            fig, ax = plt.subplots(figsize=(6, 3), facecolor="#0f172a")
            ax.set_facecolor("#1e293b")
            bars = ax.bar(years, vals, color="#6366f1", alpha=0.85, width=0.55)
            ax.tick_params(colors="#94a3b8", labelsize=8)
            ax.spines[:].set_color("#334155")
            ax.set_title("Revenue (₹ Cr)", color="#e2e8f0", fontsize=9, pad=6)
            for bar, v in zip(bars, vals):
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(vals)*0.01,
                        f"{v:,.0f}", ha="center", va="bottom", color="#94a3b8", fontsize=7)
            charts.append({"name": "Revenue", "type": "bar", "image_base64": _b64(fig)})
    except Exception as e:
        logger.debug(f"[charts] revenue failed: {e}")

    # ── Chart 3: EBITDA Margin ───────────────────────────────────────────────
    try:
        fin = t.financials
        if fin is not None and not fin.empty:
            ebit_row = fin.loc["EBITDA"].dropna() if "EBITDA" in fin.index else None
            rev_row  = fin.loc["Total Revenue"].dropna() if "Total Revenue" in fin.index else None
            if ebit_row is not None and rev_row is not None:
                common = sorted(set(ebit_row.index) & set(rev_row.index), reverse=True)[:4]
                years = [str(c)[:4] for c in common][::-1]
                margins = [float(ebit_row[c]) / float(rev_row[c]) * 100 for c in common][::-1]
                fig, ax = plt.subplots(figsize=(6, 3), facecolor="#0f172a")
                ax.set_facecolor("#1e293b")
                ax.plot(years, margins, color="#10b981", marker="o", linewidth=2, markersize=5)
                ax.fill_between(years, margins, alpha=0.12, color="#10b981")
                ax.tick_params(colors="#94a3b8", labelsize=8)
                ax.spines[:].set_color("#334155")
                ax.set_title("EBITDA Margin (%)", color="#e2e8f0", fontsize=9, pad=6)
                ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.1f}%"))
                charts.append({"name": "EBITDA Margin", "type": "line", "image_base64": _b64(fig)})
    except Exception as e:
        logger.debug(f"[charts] ebitda margin failed: {e}")

    # ── Chart 4: Net Margin ──────────────────────────────────────────────────
    try:
        fin = t.financials
        if fin is not None and not fin.empty:
            ni_row  = fin.loc["Net Income"].dropna() if "Net Income" in fin.index else None
            rev_row = fin.loc["Total Revenue"].dropna() if "Total Revenue" in fin.index else None
            if ni_row is not None and rev_row is not None:
                common = sorted(set(ni_row.index) & set(rev_row.index), reverse=True)[:4]
                years = [str(c)[:4] for c in common][::-1]
                margins = [float(ni_row[c]) / float(rev_row[c]) * 100 for c in common][::-1]
                fig, ax = plt.subplots(figsize=(6, 3), facecolor="#0f172a")
                ax.set_facecolor("#1e293b")
                colors = ["#ef4444" if m < 0 else "#f59e0b" for m in margins]
                ax.bar(years, margins, color=colors, alpha=0.85, width=0.55)
                ax.axhline(0, color="#475569", linewidth=0.8)
                ax.tick_params(colors="#94a3b8", labelsize=8)
                ax.spines[:].set_color("#334155")
                ax.set_title("Net Margin (%)", color="#e2e8f0", fontsize=9, pad=6)
                ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.1f}%"))
                charts.append({"name": "Net Margin", "type": "bar", "image_base64": _b64(fig)})
    except Exception as e:
        logger.debug(f"[charts] net margin failed: {e}")

    # ── Chart 5: Revenue Growth YoY ──────────────────────────────────────────
    try:
        fin = t.financials
        if fin is not None and not fin.empty and "Total Revenue" in fin.index:
            rev_row = fin.loc["Total Revenue"].dropna()
            cols = sorted(rev_row.index, reverse=True)[:5]
            if len(cols) >= 2:
                years = [str(c)[:4] for c in cols][::-1]
                revs  = [float(rev_row[c]) for c in cols][::-1]
                growths = [None] + [
                    (revs[i] - revs[i-1]) / abs(revs[i-1]) * 100 if revs[i-1] else 0
                    for i in range(1, len(revs))
                ]
                g_years = years[1:]
                g_vals  = growths[1:]
                fig, ax = plt.subplots(figsize=(6, 3), facecolor="#0f172a")
                ax.set_facecolor("#1e293b")
                clrs = ["#ef4444" if v < 0 else "#22d3ee" for v in g_vals]
                ax.bar(g_years, g_vals, color=clrs, alpha=0.85, width=0.55)
                ax.axhline(0, color="#475569", linewidth=0.8)
                ax.tick_params(colors="#94a3b8", labelsize=8)
                ax.spines[:].set_color("#334155")
                ax.set_title("Revenue Growth YoY (%)", color="#e2e8f0", fontsize=9, pad=6)
                ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.1f}%"))
                charts.append({"name": "Revenue Growth YoY", "type": "bar", "image_base64": _b64(fig)})
    except Exception as e:
        logger.debug(f"[charts] rev growth failed: {e}")

    # ── Chart 6: Volume (3M) ─────────────────────────────────────────────────
    try:
        hist = t.history(period="3mo")
        if not hist.empty and "Volume" in hist.columns:
            fig, ax = plt.subplots(figsize=(6, 3), facecolor="#0f172a")
            ax.set_facecolor("#1e293b")
            vol_m = hist["Volume"] / 1e6
            avg   = vol_m.rolling(20).mean()
            ax.bar(hist.index, vol_m, color="#6366f1", alpha=0.5, width=1.0)
            ax.plot(hist.index, avg, color="#f59e0b", linewidth=1.5, label="20-day avg")
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%b '%y"))
            ax.tick_params(colors="#94a3b8", labelsize=8)
            ax.spines[:].set_color("#334155")
            ax.set_title("Volume (M shares, 3M)", color="#e2e8f0", fontsize=9, pad=6)
            ax.legend(fontsize=7, facecolor="#1e293b", labelcolor="#94a3b8", framealpha=0.6)
            plt.setp(ax.get_xticklabels(), rotation=30, ha="right")
            charts.append({"name": "Volume (3M)", "type": "bar", "image_base64": _b64(fig)})
    except Exception as e:
        logger.debug(f"[charts] volume failed: {e}")

    return charts


_charts_cache: dict = {}

@api_router.get("/research/{session_id}/charts")
async def research_charts(session_id: str):
    """Return 6 financial charts as base64 PNGs."""
    try:
        session = await _load_session_for_report(session_id)
        ticker = (session.get("ticker") or "").upper()
        if not ticker:
            return {"charts": []}
        # 30-min cache per ticker
        import time as _time
        now = _time.time()
        cached = _charts_cache.get(ticker)
        if cached and now - cached[0] < 1800:
            return {"charts": cached[1]}
        import asyncio
        loop = asyncio.get_event_loop()
        charts = await loop.run_in_executor(None, _generate_charts, ticker)
        _charts_cache[ticker] = (now, charts)
        return {"charts": charts}
    except Exception as e:
        logger.error(f"[charts] {session_id}: {e}")
        return {"charts": []}


# ─────────────────────────────────────────────────────────────────────────────
# FEATURE 15 — Peer Comparison (live data)
# ─────────────────────────────────────────────────────────────────────────────

def _fetch_peer_metrics(tickers: list[str]) -> list[dict]:
    """Fetch P/E, EV/EBITDA, ROE for a list of tickers via yfinance."""
    import yfinance as yf
    rows = []
    for tk in tickers:
        try:
            ticker_yf = tk if (tk.endswith(".NS") or tk.endswith(".BO")) else tk + ".NS"
            info = yf.Ticker(ticker_yf).info or {}
            pe     = info.get("trailingPE") or info.get("forwardPE")
            ev_rev = info.get("enterpriseToEbitda")
            roe    = info.get("returnOnEquity")
            name   = info.get("shortName") or info.get("longName") or tk
            rows.append({
                "ticker":     tk,
                "name":       name[:25] if name else tk,
                "pe_fy25e":   round(float(pe), 1) if pe else None,
                "ev_ebitda":  round(float(ev_rev), 1) if ev_rev else None,
                "roe":        round(float(roe) * 100, 1) if roe else None,
            })
        except Exception as ex:
            logger.debug(f"[peers] {tk} failed: {ex}")
            rows.append({"ticker": tk, "name": tk, "pe_fy25e": None, "ev_ebitda": None, "roe": None})
    return rows


_peers_cache: dict = {}

@api_router.get("/research/{session_id}/peers")
async def research_peers(session_id: str):
    """Return peer comparison metrics."""
    try:
        session = await _load_session_for_report(session_id)
        ticker = (session.get("ticker") or "").upper()
        sector = (session.get("sector") or "").upper()
        if not ticker:
            return []
        # Determine peer list
        peer_key = None
        for k in PEER_MAP:
            if k != "DEFAULT" and k in sector:
                peer_key = k
                break
        peer_tickers = PEER_MAP.get(peer_key) if peer_key else PEER_MAP["DEFAULT"]
        peer_tickers = [p for p in peer_tickers if p.upper() != ticker.upper()][:4]
        # Include subject ticker first
        all_tickers = [ticker] + peer_tickers
        import time as _time
        now = _time.time()
        cache_key = ",".join(all_tickers)
        cached = _peers_cache.get(cache_key)
        if cached and now - cached[0] < 3600:
            return cached[1]
        import asyncio
        loop = asyncio.get_event_loop()
        rows = await loop.run_in_executor(None, _fetch_peer_metrics, all_tickers)
        _peers_cache[cache_key] = (now, rows)
        return rows
    except Exception as e:
        logger.error(f"[peers] {session_id}: {e}")
        return []


# ─────────────────────────────────────────────────────────────────────────────
# FEATURE 16 — Factor Engine scores (momentum / value / quality / macro)
# ─────────────────────────────────────────────────────────────────────────────

@api_router.get("/research/{session_id}/factors")
async def research_factors(session_id: str):
    """Return momentum, value, quality, macro factor scores (0-100)."""
    try:
        session = await _load_session_for_report(session_id)
        scoring = session.get("scoring") or {}
        # Map scoring dimensions → 4 factor buckets
        fs  = float(scoring.get("financial_strength",      50) or 50)
        gq  = float(scoring.get("growth_quality",          50) or 50)
        va  = float(scoring.get("valuation_attractiveness",50) or 50)
        risk= float(scoring.get("risk",                    50) or 50)
        mp  = float(scoring.get("market_positioning",      50) or 50)
        return {
            "factors": {
                "momentum": round(min(100, max(0, (gq * 0.6 + mp * 0.4))), 1),
                "value":    round(min(100, max(0, va)), 1),
                "quality":  round(min(100, max(0, (fs * 0.6 + mp * 0.4))), 1),
                "macro":    round(min(100, max(0, 100 - risk)), 1),
            }
        }
    except Exception as e:
        logger.error(f"[factors] {session_id}: {e}")
        return {"factors": {"momentum": 50, "value": 50, "quality": 50, "macro": 50}}


# =====================================================================
# MOUNT ROUTER — LAST, ALWAYS
# =====================================================================

app.include_router(api_router)

# Global exception handler — last line of defense
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.error(f"unhandled exception on {request.url.path}: {exc}\n{traceback.format_exc()}")
    return JSONResponse(
        status_code=200,  # never 500 to the frontend
        content={
            "error": str(exc),
            "path": str(request.url.path),
            "_fallback": True,
        },
    )

logger.info(f"Market Pulse hardened backend ready | version={VERSION} | mongo={'up' if _get_db() is not None else 'down'}")
